from datetime import date
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
from django.db.models import Q
from django.db.models import Sum, Count
import xlwt
from django.views.generic import ListView, DetailView, View

from docxtpl import DocxTemplate, InlineImage #r

# Create your views here.
from openpyxl.styles.builtins import output

from .forms import mota_cv_form, Danhgia_nlform
from .forms import nang_luc_cv_form, Congviec_nangluc_form, Danhgia_nangluc_form, Khung_nangluc, Danhgia_nangluc_up_formQL
from .models import Congviec_nangluc, Danhgia_nluc, tong_nl
from mota_cv.models import Mota_Cv, Mota_Cv7
from enroll.models import Nhan_vien
from nhansu.models import Nang_luc_2

import json

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

# Create your views here.
from django.views.decorators.csrf import csrf_exempt

#-------------------

#-------------------Khung update năng lực---------Danhgia_nangluc_quanly.html-----------------------------------------------------
def khungnangluc_view(request): #Danhgia_nangluc_quanly.html
    chucdanh_CV = request.GET.get('chucdanh_CV')
    form = Congviec_nangluc_form(request.POST or None)
    danhgia_nangluc_list = Congviec_nangluc.objects.all()[1:10]
    context = {'students': danhgia_nangluc_list, 'form': form, }

    if request.method == 'POST':
        name = form['name'].value()
        chucdanh_CV = form['chucdanh_CV'].value()

        if (name != ''):
            danhgia_nangluc_list = Congviec_nangluc.objects.filter(
                name__icontains=form['name'].value(),
            )
        if (chucdanh_CV != ''):
            # queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV,nangluc_cv_id = nangluc_cv)
            danhgia_nangluc_list = Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV)

    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)
    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form}

    Diem_kq = Congviec_nangluc.objects.aggregate(Count('id'))

    from django.db.models import Sum
    total_mucQT = (danhgia_nangluc_list.aggregate(total=Sum('Muc_quantrong_nluc', field="Muc_quantrong_nluc"))['total'])
    total_mucTT = (danhgia_nangluc_list.aggregate(total=Sum('Muc_thanhthao_nluc', field="Muc_thanhthao_nluc*10"))['total'])
    total_diemTC = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tieuchuan', field="Diem_tieuchuan"))['total'])

    if form['Xuất_Excel'].value() == True:
            responese = HttpResponse(content_type='application/ms-excel')
            # responese['Content-Disposition'] = 'attachment; filename=Khung_Nang_luc'+'  '+'.xls'
            responese['Content-Disposition'] = 'attachment; filename=' + chucdanh_CV + '_KhungNL.xls'

            wb = xlwt.Workbook(encoding='utf-8')
            # wb.set_paper(9)
            # wb.repeat_rows(10, 10)

            ws = wb.add_sheet(str(chucdanh_CV))

            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )
            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 5
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 5
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 5
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'PHIẾU TIÊU CHUẨN NĂNG LỰC', TIEUDEM)

            rows1 = danhgia_nangluc_list.values_list('chucdanh_CV__Ten_Nhom_CV')[1]
            rows3 = danhgia_nangluc_list.values_list('chucdanh_CV__bo_phan__ten_bp')[1]
            rows4 = danhgia_nangluc_list.values_list('chucdanh_CV__id')[1]

            ws.write(5, 2, "Chức danh công việc:", TIEUDE2)
            top_row8 = 5
            bottom_row8 = 5
            left_column8 = 3
            right_column8 = 5
            ws.write_merge(top_row8, bottom_row8, left_column8, right_column8,   rows1, TIEUDE2) #cHỨC DANH

            ws.write(6, 2, "Đơn vị:", TIEUDE2)

            top_rowdv = 6
            bottom_rowdv = 6
            left_columndv = 3
            right_columndv = 5
            ws.write_merge(top_rowdv, bottom_rowdv, left_columndv, right_columndv,    rows3, TIEUDE2)
            ws.write(7, 2, "Mã chức danh:", TIEUDE2)
            # ws.write(7,2,  rows4, TIEUDE2)
            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 2000
            ws.col(1).width = 2000
            ws.col(2).width = 10000
            ws.col(3).width = 4000
            ws.col(4).width = 4000
            ws.col(5).width = 4000
            ws.col(6).width = 6000

            columns = ['STT', 'Mã ', 'Tên năng lực', ' Mức độ quan trọng', 'Mức độ thành thạo', 'Điểm tiêu chuẩn', ]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = danhgia_nangluc_list.values_list('stt', 'nangluc_cv__ma_nangluc', 'nangluc_cv__name', 'Muc_quantrong_nluc',
                                        'Muc_thanhthao_nluc', 'Diem_tieuchuan', )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )

            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'TỔNG ĐIỂM', TABLE_HEADER)
            ws.write(row_num + 1, 3, total_mucQT, TABLE_HEADER)
            ws.write(row_num + 1, 4, total_mucTT, TABLE_HEADER)
            ws.write(row_num + 1, 5, total_diemTC, TABLE_HEADER)


            ws.write(row_num + 5, 1, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese


    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': danhgia_nangluc_list,
               "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form,

               'form': form, 'total_mucQT': total_mucQT,
               'total_mucTT': total_mucTT,
               'total_diemTC':total_diemTC,
               }

    return render(request,"nangluc/Khung_nangluc_moi.html",context)

@csrf_exempt#Danhgia_nangluc_aab.html
def Insertkhung_nangluc(request):
    nangluc_cv=request.POST.get("nangluc_cv")
    #Quanly_danhgia=request.POST.get("Quanly_danhgia")

    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    Muc_thanhthao_nluc=request.POST.get("Muc_thanhthao_nluc")

    try:
        student=Congviec_nangluc(nangluc_cv=nangluc_cv, tu_danhgia_dapung=tu_danhgia_dapung, Muc_thanhthao_nluc=Muc_thanhthao_nluc,)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)


@csrf_exempt#Danhgia_nangluc_aab.html
def update_all_k(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Congviec_nangluc.objects.get(id=dic_single['id'])
           # student.nangluc_cv=dic_single['nangluc_cv']
            student.Muc_quantrong_nluc=dic_single['Muc_quantrong_nluc']
            student.Muc_thanhthao_nluc=dic_single['Muc_thanhthao_nluc']


            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt#Danhgia_nangluc_aab.html
def delete_data_k(request):
    id=request.POST.get("id")
    try:
        student=Congviec_nangluc.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xóa không được"}
        return JsonResponse(stuent_data,safe=False)

#-------------------Kếtthúc Khunng  nang lực----------------------------------------------------------






















































@login_required
#Khungnangluc_list_2.htm
def Congviec_nangluc_list_2(request):
    form = mota_cv_form(request.POST or None)
    chucdanh_CV = request.GET.get('chucdanh_CV')
    print("chucdanh_CV:" 'chucdanh_CV')

    if chucdanh_CV == None:
        Congviec_nangluc_list_2 = Congviec_nangluc.objects.all()[1:7]
    else:
        Congviec_nangluc_list_2 = Congviec_nangluc.objects.filter(chucdanh_CV__Ten_vitri_full=chucdanh_CV)

    chucdanh_CV7 = Mota_Cv7.objects.filter(don_vi_id=13)

    context = {'chucdanh_CV7': chucdanh_CV7, 'list_nangluc': Congviec_nangluc_list_2, 'form': form}

    return render(request, 'nangluc/Khungnangluc_list_2.html', context)

# ----------------------------------------
def update_Congviec_nangluc(request, id):
    if request.method == 'POST':
        dv = Congviec_nangluc.objects.get(pk=id)
        fmdv = Congviec_nangluc_form(request.POST, instance=dv)
        if fmdv.is_valid():
            fmdv.save()
    else:
        dv = Congviec_nangluc.objects.get(pk=id)
        fmdv = Congviec_nangluc_form(instance=dv)
    return render(request, 'nangluc/Khung_nangluc_up_CN.html', {'form': fmdv})

def del_Congviec_nangluc(request, id):
    dv = Congviec_nangluc.objects.get(pk=id)
    if request.method == 'POST':
        dv.delete()
        return HttpResponseRedirect('/cvnl/')
    return render(request, 'nangluc/Khung_nangluc_delete.html')
# ---------------------------------THIẾT LẬP KHUNG NĂNMG LỤC:--------------
@login_required
def add_khung_nangluc(request):
    ten_nhom_nangluc = Mota_Cv7.objects.all()
    if request.method == 'POST':
        data = request.POST
        Nang_luc_2s = Nang_luc_2.objects.filter(loai_nang_luc_id=6)  # Năng lực: chung=1; quản lý =3; chuyên môn =2
        # Mota_Cv7s= Mota_Cv7.objects.all()
        Mota_Cv7s = Mota_Cv7.objects.filter(Loai_laodong_id=2)
        # Mota_Cv7s= Mota_Cv7.objects.filter (Q(Loai_laodong_id=3)|Q(Loai_laodong_id=4)|Q(Loai_laodong_id=5)|Q(Loai_laodong_id=6)|Q(Loai_laodong_id=7))
        print('data:', data)
        print('Tên công việc', Nang_luc_2s)

        for nangluc_cv in Nang_luc_2s:
            for chucdanh_CV in Mota_Cv7s:
                khoa_hoc = Congviec_nangluc.objects.update_or_create(
                    name=data['name'],
                    nangluc_cv=nangluc_cv,
                    Muc_quantrong_nluc=data['Muc_quantrong_nluc'],
                    Muc_thanhthao_nluc=data['Muc_thanhthao_nluc'],
                    Diem_tieuchuan=data['Diem_tieuchuan'],
                    chucdanh_CV=chucdanh_CV,
                )
        return redirect('list_nangluc_2')
    context = {'ten_nhom_nangluc': ten_nhom_nangluc}
    return render(request, 'nangluc/Khung_nangluc_add.html')


# ------------Khung năng lực CM
def add_khung_nanglucchuyenmon(request):
    chucdanh_CV = Mota_Cv7.objects.all()
    nangluc_cv = Nang_luc_2.objects.all()
    if request.method == 'POST':
        data = request.POST
        if data['chucdanh_CV'] != 'none':
            chucdanh_CV = Mota_Cv7.objects.get(id=data['ten_vitri'])
        if data['ten_nhom_nangluc'] != 'none':
            ten_nhom_nangluc = Nang_luc_2.objects.get(id=data['ten_nhom_nangluc'])

            khoa_hoc = Congviec_nangluc.objects.create(
                name=data['name'],
                nangluc_cv=nangluc_cv,
                Muc_quantrong_nluc=data['Muc_quantrong_nluc'],
                Muc_thanhthao_nluc=data['Muc_thanhthao_nluc'],
                Diem_tieuchuan=data['Diem_tieuchuan'],
                chucdanh_CV=chucdanh_CV,
            )
            context = {'nangluc_cv2': nangluc_cv, 'chucdanh_CV_s': chucdanh_CV}

    return render(request, 'nangluc/Khung_nanglucCM_add.html')


# ---------------------------------END thiết lập khung N8ang lực:--------------



@login_required
def Khung_nangluc_list(request):
    form = Khung_nangluc(request.POST or None)
    queryset = Congviec_nangluc.objects.all()[0:10]
    context = {'queryset': queryset, 'form': form, }

    if request.method == 'POST':
        nangluc_cv = form['nangluc_cv'].value()
        chucdanh_CV = form['chucdanh_CV'].value()
        name = form['name'].value()

        if (name != ''):
            queryset = Congviec_nangluc.objects.filter(
                name__icontains=form['name'].value(),
            )
        if (chucdanh_CV != ''):
            # queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV,nangluc_cv_id = nangluc_cv)
            queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV)


        context = {'queryset': queryset, 'form': form, }


        total_muc_qt = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Muc_quantrong_nluc', field="Muc_quantrong_nluc"))['total'])
        total_muc_tt = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Muc_thanhthao_nluc', field="Muc_thanhthao_nluc"))['total'])
        Diem_tieuchuan = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Diem_tieuchuan', field="Diem_tieuchuan"))['total'])


        if form['Xuất_Excel'].value() == True:


            responese = HttpResponse(content_type='application/ms-excel')
            # responese['Content-Disposition'] = 'attachment; filename=Khung_Nang_luc'+'  '+'.xls'
            responese['Content-Disposition'] = 'attachment; filename=' + chucdanh_CV + '_KhungNL.xls'

            wb = xlwt.Workbook(encoding='utf-8')
            # wb.set_paper(9)
            # wb.repeat_rows(10, 10)

            ws = wb.add_sheet(str(chucdanh_CV))

            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )
            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 5
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 5
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 5
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'PHIẾU TIÊU CHUẨN NĂNG LỰC', TIEUDEM)

            rows1 = queryset.values_list('chucdanh_CV__Ten_Nhom_CV')[1]
            rows3 = queryset.values_list('chucdanh_CV__bo_phan__ten_bp')[1]
            rows4 = queryset.values_list('chucdanh_CV__id')[1]

            ws.write(5, 2, "Chức danh công việc:", TIEUDE2)
            top_row8 = 5
            bottom_row8 = 5
            left_column8 = 3
            right_column8 = 5
            ws.write_merge(top_row8, bottom_row8, left_column8, right_column8,   rows1, TIEUDE2) #cHỨC DANH

            ws.write(6, 2, "Đơn vị:", TIEUDE2)

            top_rowdv = 6
            bottom_rowdv = 6
            left_columndv = 3
            right_columndv = 5
            ws.write_merge(top_rowdv, bottom_rowdv, left_columndv, right_columndv,    rows3, TIEUDE2)
            ws.write(7, 2, "Mã chức danh:", TIEUDE2)
            # ws.write(7,2,  rows4, TIEUDE2)
            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 2000
            ws.col(1).width = 2000
            ws.col(2).width = 10000
            ws.col(3).width = 4000
            ws.col(4).width = 4000
            ws.col(5).width = 4000
            ws.col(6).width = 6000

            columns = ['STT', 'Mã ', 'Tên năng lực', ' Mức độ quan trọng', 'Mức độ thành thạo', 'Điểm tiêu chuẩn', ]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = queryset.values_list('stt', 'nangluc_cv__ma_nangluc', 'nangluc_cv__name', 'Muc_quantrong_nluc',
                                        'Muc_thanhthao_nluc', 'Diem_tieuchuan', )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )
            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'TỔNG ĐIỂM', TABLE_HEADER)
            ws.write(row_num + 1, 3, total_muc_qt, TABLE_HEADER)
            ws.write(row_num + 1, 4, total_muc_tt, TABLE_HEADER)
            ws.write(row_num + 1, 5, Diem_tieuchuan, TABLE_HEADER)
            ws.write(row_num + 5, 1, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese
        total_queryset = queryset.count()
        context = {'queryset': queryset, 'form': form}
    return render(request, 'nangluc/Khung_nangluc_list.html', context)

##<<<<<<<<<<<<: Kết thúc Chương trình Hiện thị List+ Excel NHÂN VIÊN

# -----<<<

@login_required
#Danhgia_nangluc2.htm
def Danhgia_nangluc_list2(request):
    Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    form = Danhgia_nlform(request.POST or None)
    if Nhanvien_dg_nangluc == None:
        danhgia_nangluc_list = Danhgia_nluc.objects.all()[1:50]
    else:

        danhgia_nangluc_list = Danhgia_nluc.objects.filter(Nhanvien_dg_nangluc__ho_lot_thuong_dung=Nhanvien_dg_nangluc)
    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list2': danhgia_nangluc_list, 'form': form}
    Diem_kq = Danhgia_nluc.objects.aggregate(Count('Ketqua_danhgia'))

    from django.db.models import Sum
    total_tu_dg = (danhgia_nangluc_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (danhgia_nangluc_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_nangluc_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])

    total_mucqt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_quantrong_nluc',
                                                  field="TenNangluc_congviec__Muc_quantrong_nluc"))['total'])

    total_tt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_thanhthao_nluc',
                                                  field="TenNangluc_congviec.Muc_thanhthao_nluc"))['total'])

    total_diemchuan = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tieuchuan',
                                                   field="Diem_tieuchuan"))['total'])

    total_diemcdat = (danhgia_nangluc_list.aggregate(total=Sum('Diem_dat',
                                                   field="Diem_dat*1"))['total'])
    ketqua_nangluc = (total_diemchuan / total_diemcdat)

    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_NL_'+'.xls'
        #responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_nangluc + '_KhungNL.xls'

        #responese['Content-Disposition'] = 'attachment; filename='+ Nhanvien_dg_nangluc + '.xls'
       # responese['Content-Disposition'] = 'attachment; filename=str(Nhanvien_dg_nangluc)'+'d.xls'

#===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 3
        right_column9 = 6
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ NĂNG LỰC NHÂN SỰ', TIEUDEM)

        rows1 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1]
        rows2 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 1800
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'Tên năng lực', 'Mức độ quan trọng', 'Mức độ Thành thạo', 'Điểm tiêu chuẩn',' Tự đánh giá',
                   'Cấp trên đánh giá', 'Kết quả thống nhất','Điểm ĐG',  'Kết quả']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_nangluc_list.values_list('TenNangluc_congviec__stt', 'TenNangluc_congviec__nangluc_cv__ma_nangluc',
                                                'TenNangluc_congviec__nangluc_cv__name',
                                                'TenNangluc_congviec__Muc_quantrong_nluc',
                                                'TenNangluc_congviec__Muc_thanhthao_nluc',

                                                'TenNangluc_congviec__Diem_tieuchuan',
                                                'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Diem_dat', 'Kha_nang_dapung')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 3, total_mucqt, TABLE_HEADER)
        ws.write(row_num + 1, 4, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 5, total_diemchuan, TABLE_HEADER)
        ws.write(row_num + 1, 6, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 7, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 8, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_diemchuan, TABLE_HEADER)
        ws.write(row_num + 1, 10, tam1, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ... năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    if form['Xuất_Word'].value() == True:
        doc = DocxTemplate("word_template/M_danhgiangluc.docx")
        queryset5 = {
                  "Ho_ten":Nhanvien_dg_nangluc,
                  "Chuc_danh":(danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV'))[1] ,

                  "Don_vi":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1] ,
                  'total_diemchuan': total_diemchuan,
                  'total_diemcdat': total_diemcdat,
                  'ketqua_nangluc' :ketqua_nangluc,

                    }
        doc.render(queryset5)
        doc.save('thu_word/NL_'+ Nhanvien_dg_nangluc + '.docx')

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list2': danhgia_nangluc_list,
               'total_mucqt': total_mucqt, 'total_tt': total_tt,
               'total_diemchuan': total_diemchuan,
               'total_diemcdat': total_diemcdat,
               'ketqua_nangluc': ketqua_nangluc,

               'form': form, 'total_tu_dg': total_tu_dg, 'total_ql': total_ql}
    return render(request, 'nangluc/Danhgia_nangluc2.html', context)


def Danhgia_nangluc_list(request):
    Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    form = Danhgia_nangluc_form(request.POST or None)
    danhgia_nangluc_list = Danhgia_nluc.objects.all()[1:10]
    context = {'danhgia_nangluc_list1': danhgia_nangluc_list, 'form': form, }
    if request.method == 'POST':
        Landanhgia_nagluc = form['Landanhgia_nagluc'].value()
        Nhanvien_dg_nangluc = form['Nhanvien_dg_nangluc'].value()
        TenNangluc_congviec = form['TenNangluc_congviec'].value()

        if (Landanhgia_nagluc != ''):
            danhgia_nangluc_list = Danhgia_nluc.objects.filter(
                Landanhgia_nagluc__icontains=form['Landanhgia_nagluc'].value(),
            )
        if (Nhanvien_dg_nangluc != ''):
            # queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV,nangluc_cv_id = nangluc_cv)
            danhgia_nangluc_list = Danhgia_nluc.objects.filter(Nhanvien_dg_nangluc_id=Nhanvien_dg_nangluc)

    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': danhgia_nangluc_list, 'form': form}
    Diem_kq = Danhgia_nluc.objects.aggregate(Count('Ketqua_danhgia'))

    from django.db.models import Sum
    total_tu_dg = (danhgia_nangluc_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (danhgia_nangluc_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_nangluc_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])


    total_mucqt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_quantrong_nluc',
                                                  field="TenNangluc_congviec__Muc_quantrong_nluc"))['total'])

    total_tt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_thanhthao_nluc',
                                                  field="TenNangluc_congviec.Muc_thanhthao_nluc"))['total'])

    total_diemchuan = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tieuchuan',
                                                   field="Diem_tieuchuan"))['total'])

    total_diemcdat = (danhgia_nangluc_list.aggregate(total=Sum('Diem_dat',
                                                   field="Diem_dat*Diem_tieuchuan"))['total'])

    total_Diem_tu_danhgia = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tu_danhgia',
                                                   field="Diem_dat*1"))['total'])


    if  total_diemchuan  and total_Diem_tu_danhgia :
        ketqua_nangluc =  (total_Diem_tu_danhgia/total_diemchuan )
    else:
        ketqua_nangluc =0

    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_NL_'+'.xls'
        #responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_nangluc + '_KhungNL.xls'
        #responese['Content-Disposition'] = 'attachment; filename='+ Nhanvien_dg_nangluc + '.xls'
       # responese['Content-Disposition'] = 'attachment; filename=str(Nhanvien_dg_nangluc)'+'d.xls'
#===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        #ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        ws = wb.add_sheet('ĐG')
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ NĂNG LỰC NHÂN SỰ', TIEUDEM)

#  --------------------------------------------
        rows1 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1]
        rows2 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 1800
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'Tên năng lực', 'Mức độ quan trọng', 'Mức độ Thành thạo', 'Điểm tiêu chuẩn',' Tự đánh giá',
                   'Cấp trên đánh giá', 'Kết quả thống nhất','Điểm ĐG',  'Kết quả']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_nangluc_list.values_list('TenNangluc_congviec__stt', 'TenNangluc_congviec__nangluc_cv__ma_nangluc',
                                                'TenNangluc_congviec__nangluc_cv__name',
                                                'TenNangluc_congviec__Muc_quantrong_nluc',
                                                'TenNangluc_congviec__Muc_thanhthao_nluc',

                                                'TenNangluc_congviec__Diem_tieuchuan',
                                                'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Diem_dat', 'Kha_nang_dapung')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 3, total_mucqt, TABLE_HEADER)
        ws.write(row_num + 1, 4, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 5, total_diemchuan, TABLE_HEADER)

        ws.write(row_num + 1, 6, total_tu_dg, TABLE_HEADER)# Tự đánh giá
        ws.write(row_num + 1, 7, total_ql, TABLE_HEADER) #Cấp trên đánh giá
        ws.write(row_num + 1, 8, total_chung, TABLE_HEADER) #Kết quả thống nhất
        ws.write(row_num + 1, 9, total_diemcdat, TABLE_HEADER) #Điểm đánh giá
        ws.write(row_num + 1, 10, ketqua_nangluc, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ...năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ...năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    if form['Xuất_Word'].value() == True:

        doc = DocxTemplate("word_template/M_danhgiangluc.docx")
        queryset5 = {
                  "Ho_ten":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1] ,
                  "Chuc_danh":(danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV'))[1] ,

                  "Don_vi":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1] ,
                  'total_diemchuan': total_diemchuan,
                  'total_diemcdat': total_diemcdat,
                  'tiledat' :98,

                   "productss": danhgia_nangluc_list,

                    }
        doc.render(queryset5)
        doc.save('thu_word/NL_'+ Nhanvien_dg_nangluc + '.docx')

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': danhgia_nangluc_list,
               'total_mucqt': total_mucqt, 'total_tt': total_tt,
               'total_diemchuan': total_diemchuan,
               'total_diemcdat': total_diemcdat,
               'form': form, 'total_tu_dg': total_tu_dg,
               'total_ql': total_ql,
               'total_chung':total_chung,

                "productss": danhgia_nangluc_list,

               'ketqua_nangluc': ketqua_nangluc,}

    return render(request, 'nangluc/Danhgia_nangluc_chitiet.html', context)



# ---------------------------------
@login_required
def add_danhgia_nangluc(request):
    # Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    Nhanvien_danhgia = Nhan_vien.objects.all()
    if request.method == 'POST':
        data = request.POST
        # Nhan_viens = Nhan_vien.objects.filter(Q(vitri_CV_id=1)|Q(vitri_CV_id=2))
        Nhan_viens = Nhan_vien.objects.filter()
        Congviec_nanglucs = Congviec_nangluc.objects.filter()
        khoa_ = Danhgia_nluc.objects
        # Congviec_nanglucs = Congviec_nangluc.objects.all()
     #   print('Nhập dự liệu ở form:', data)
    #    print('Tên Nhân viên', Nhan_viens)
        for Nhanvien_dg_nangluc in Nhan_viens:
            for TenNangluc_congviec in Congviec_nanglucs:
                if Nhanvien_dg_nangluc.vitri_CV_id == TenNangluc_congviec.chucdanh_CV_id:
                    khoa_hoc = Danhgia_nluc.objects.update_or_create(
                        Landanhgia_nagluc=data['Landanhgia_nagluc'],
                        Nhanvien_dg_nangluc=Nhanvien_dg_nangluc,
                        TenNangluc_congviec=TenNangluc_congviec,

                        tu_danhgia_dapung = TenNangluc_congviec.Muc_thanhthao_nluc,
                        Quanly_danhgia = TenNangluc_congviec.Muc_thanhthao_nluc, # lấy mức điểm thống nhất bằng điểm Thành thạochuẩn
                        Ketqua_danhgia = TenNangluc_congviec.Muc_thanhthao_nluc, # lấy mức điểm thống nhất bằng điểm Thành thạochuẩn

                        Diem_tieuchuan = TenNangluc_congviec.Diem_tieuchuan,
                        Diem_dat = TenNangluc_congviec.Diem_tieuchuan, # Tạm lấy điểm đạt băng điểm ti6u chuẩn
                        )
                else:
                    continue
        return redirect('list_nangluc_2')
    context = {'Nhanvien_danhgias': Nhanvien_danhgia}

    return render(request, 'nangluc/Danhgia_nangluc_add.html')



def Danhgia_nangluc_update_cannhan(request, id):
    if request.method == 'POST':
        up_dgnl = Danhgia_nluc.objects.get(pk=id)
        fmdg = Danhgia_nangluc_up_formQL(request.POST, instance=up_dgnl)
        if fmdg.is_valid():
            fmdg.save()
            messages.success(request, 'Dữ liệu cập nhật')
    else:
        dv = Danhgia_nluc.objects.get(pk=id)
        fmdg = Danhgia_nangluc_up_formQL(instance=dv)
    return render(request, 'nangluc/Danhgia_up_ca_nhan.html', {'form': fmdg})



def Danhgia_nangluc_up_QL(request, id):
    if request.method == 'POST':
        up_dgnl = Danhgia_nluc.objects.get(pk=id)
        fmdgql = Danhgia_nangluc_up_formQL(request.POST, instance=up_dgnl)
        if fmdgql.is_valid():
            fmdgql.save()
            messages.success(request, 'Dữ liệu cập nhật')
    else:
        dv = Danhgia_nluc.objects.get(pk=id)
        fmdgql = Danhgia_nangluc_up_formQL(instance=dv)
        return render(request, 'nangluc/Danhgia_up.html', {'form': fmdgql})







# >>>>>>>>>>>: Chương trình HIỂN THỊ List+ Xuât excel nhân viên:
def index_nl(request):
    pro = Danhgia_nluc.objects.filter(Nhanvien_dg_nangluc_id=177)
    if request.method == 'POST':
        form = Danhgia_nangluc_form(request.POST)
        if form.is_valid():
            form.save()
            return redirect('index')
    else:
        form = Danhgia_nangluc_form()
    context = {
        "productss": pro,
        "form": form
    }
    return render(request, 'partials/index_nl.html', context)

class OrderSummaryView1(LoginRequiredMixin, View):
    def get(self, *args, **kwargs):
        try:
            gacon = tong_nl.objects.get(user=self.request.user, ordered=False)
            context = {
                'object': gacon
            }
            return render(self.request, 'nangluc/order_summary1.html', context)
        except ObjectDoesNotExist:
            messages.warning(self.request, "Bạn chưa chọn năng lực")
            return redirect("danhgia/")
# ---------------------------------
@login_required
def add_ketqua_nangluc(request):
    Nhanvien__nanglucs = Nhan_vien.objects.all()
    if request.method == 'POST':
        data = request.POST
        Nhanvien_dg_nanglucs = Nhan_vien.objects.filter()
        # Mota_Cv7s= Mota_Cv7.objects.filter (Q(Loai_laodong_id=3)|Q(Loai_laodong_id=4)|Q(Loai_laodong_id=5)|Q(Loai_laodong_id=6)|Q(Loai_laodong_id=7))
        print('data:', data)
        print('Tên công việc', Nhanvien_dg_nanglucs)
        for Nhanvienb in Nhanvien_dg_nanglucs:
            khoa_hoc = tong_nl.objects.create(
                Nhanvien_dg_nangluc=Nhanvienb,
                   )
        return redirect('Danhgia_nangluc_list2')
    context = {'ten_nhom_nangluc': Nhanvien__nanglucs}
    return render(request, 'nangluc/Tong_nangluc_add.html')

#-----------------------------------------
#----------------
@login_required
def Danhgia_nangluc_list_intatca(request):
    Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    form = Danhgia_nlform(request.POST or None)
    if Nhanvien_dg_nangluc == None:
        danhgia_nangluc_list_F = Danhgia_nluc.objects.all()[1:70]
    else:
        danhgia_nangluc_list_F = Danhgia_nluc.objects.filter(Nhanvien_dg_nangluc__ho_lot_thuong_dung=Nhanvien_dg_nangluc)


    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list2': danhgia_nangluc_list_F, 'form': form}
    Diem_kq = Danhgia_nluc.objects.aggregate(Count('Ketqua_danhgia'))

    from django.db.models import Sum
    total_tu_dg = (danhgia_nangluc_list_F.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (danhgia_nangluc_list_F.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])

    total_mucqt = (danhgia_nangluc_list_F.aggregate(total=Sum('TenNangluc_congviec__Muc_quantrong_nluc',
                                                  field="TenNangluc_congviec__Muc_quantrong_nluc"))['total'])

    total_tt = (danhgia_nangluc_list_F.aggregate(total=Sum('TenNangluc_congviec__Muc_thanhthao_nluc',
                                                  field="TenNangluc_congviec.Muc_thanhthao_nluc"))['total'])

    total_diemchuan = (danhgia_nangluc_list_F.aggregate(total=Sum('Diem_tieuchuan',
                                                   field="Diem_tieuchuan"))['total'])

    total_diemcdat = (danhgia_nangluc_list_F.aggregate(total=Sum('Diem_dat',
                                                   field="Diem_dat*1"))['total'])

    #ketqua_nangluc = total_diemcdat / total_diemchuan

    danhgia_nangluc_list = Danhgia_nluc.objects.all()


    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=NS_DG' + '.xls'
        #responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_nangluc + '_KhungNL.xls'

        #responese['Content-Disposition'] = 'attachment; filename='+ Nhanvien_dg_nangluc + '.xls'
       # responese['Content-Disposition'] = 'attachment; filename=str(Nhanvien_dg_nangluc)'+'d.xls'

#===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 12
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 12
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 12
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ NĂNG LỰC NHÂN SỰ', TIEUDEM)

        top_rowD = 5
        bottom_rowD = 5
        left_columnD = 0
        right_columnD = 12
        ws.write_merge(top_rowD, bottom_rowD, left_columnD, right_columnD,  "TẤT CẢ NHÂN SỰ TRONG ĐƠN VỊ", TIEUDEM)


        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 8500
        ws.col(2).width = 1200
        ws.col(3).width = 8500
        ws.col(4).width = 2000
        ws.col(5).width = 8500
        ws.col(6).width = 1800
        ws.col(7).width = 2000
        ws.col(8).width = 1900
        ws.col(9).width = 1900
        ws.col(10).width = 2000
        ws.col(11).width = 2000
        ws.col(12).width = 1800
        # ------------------
        columns = ['Số TT', 'Tên Nhân sự', "ID",'Đơn vị','Tên năng lực','Mã NL', 'Mức độ quan trọng', 'Mức độ Thành thạo', 'Điểm tiêu chuẩn',' Tự đánh giá',
                   'Cấp trên đánh giá', 'Kết quả thống nhất','Điểm ĐG',  'Kết quả']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False

        rows = danhgia_nangluc_list.values_list('TenNangluc_congviec__stt', 'Nhanvien_dg_nangluc__ho_lot_thuong_dung',
                                                'Nhanvien_dg_nangluc',
                                                'Nhanvien_dg_nangluc__don_vi__Ten_DV',
                                                'TenNangluc_congviec__nangluc_cv__ma_nangluc',
                                                'TenNangluc_congviec__nangluc_cv__name',
                                                'TenNangluc_congviec__Muc_quantrong_nluc',
                                                'TenNangluc_congviec__Muc_thanhthao_nluc',
                                                'TenNangluc_congviec__Diem_tieuchuan',
                                                'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Diem_dat', 'Kha_nang_dapung')

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 3, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 4, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 5, tam1, TABLE_HEADER)

        ws.write(row_num + 1, 6, total_mucqt, TABLE_HEADER)
        ws.write(row_num + 1, 7, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 8, total_diemchuan, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 10, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 11, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 12, total_diemchuan, TABLE_HEADER)
        ws.write(row_num + 1, 13, tam1, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ... năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    if form['Báo_cáo_TH'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=BC' + '.xls'
 #===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )

        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 3
        right_column7 = 8
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 3
        right_column9 = 8
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)
#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 8
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU TỔNG HỢP KẾT QUẢ ĐÁNH GIÁ NĂNG LỰC', TIEUDEM)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 8500
        ws.col(2).width = 8500
        ws.col(3).width = 8500

        ws.col(4).width = 4000

        ws.col(5).width = 3500
        ws.col(6).width = 3500
        ws.col(7).width = 2000
        ws.col(8).width = 2000

        # ------------------
        columns = ['Số TT', 'Tên Nhân sự', "Chức danh",'Đơn vị','Người đánh giá', 'Điểm tiêu chuẩn','Điểm đạt',  'Tỉ lệ đạt', 'Ghi chú']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = chucdanh_CV2.values_list('id','ho_lot_thuong_dung', 'vitri_CV__Ten_Nhom_CV', 'don_vi__Ten_DV', )

        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 3, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 4, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 5, tam1, TABLE_HEADER)

        ws.write(row_num + 1, 6, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 7, tam1, TABLE_HEADER)
        ws.write(row_num + 1, 8, tam1, TABLE_HEADER)


        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người xem xét', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ... năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người  lập', TABLE_HEADER2)
        wb.save(responese)
        return responese

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list2': danhgia_nangluc_list_F,
               'total_mucqt': total_mucqt, 'total_tt': total_tt,
               'total_diemchuan': total_diemchuan,
               'total_diemcdat': total_diemcdat,
            #   'ketqua_nangluc': ketqua_nangluc,
               'form': form, 'total_tu_dg': total_tu_dg, 'total_ql': total_ql}

    return render(request, 'nangluc/Danhgia_nangluc_intatca.html', context)

@login_required
def Khung_nangluc_intatca(request):
    form = Khung_nangluc(request.POST or None)
    queryset_f = Congviec_nangluc.objects.all()[1:50]
    context = {'queryset': queryset_f, 'form': form, }
    if request.method == 'POST':
        nangluc_cv = form['nangluc_cv'].value()
        chucdanh_CV = form['chucdanh_CV'].value()
        name = form['name'].value()

        if (name != ''):
            queryset_f = Congviec_nangluc.objects.filter(
                name__icontains=form['name'].value(),
            )
        if (chucdanh_CV != ''):
            # queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV,nangluc_cv_id = nangluc_cv)
            queryset_f = queryset_f.filter(chucdanh_CV_id=chucdanh_CV)

          #  print('chức danh công việc:''queryset')
        context = {'queryset': queryset_f, 'form': form, }

        queryset = Congviec_nangluc.objects.all()

        total_muc_qt = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Muc_quantrong_nluc', field="Muc_quantrong_nluc"))['total'])
        total_muc_tt = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Muc_thanhthao_nluc', field="Muc_thanhthao_nluc"))['total'])
        Diem_tieuchuan = (Congviec_nangluc.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Diem_tieuchuan', field="Diem_tieuchuan"))['total'])

        if form['Xuất_Excel'].value() == True:


            responese = HttpResponse(content_type='application/ms-excel')
            # responese['Content-Disposition'] = 'attachment; filename=Khung_Nang_luc'+'  '+'.xls'
            responese['Content-Disposition'] = 'attachment; filename=' + '_Tatca_KhungNL.xls'

            wb = xlwt.Workbook(encoding='utf-8')
            # wb.set_paper(9)
            # wb.repeat_rows(10, 10)
            ws = wb.add_sheet(str(chucdanh_CV))
            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )

            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 8
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 8
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 8
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'PHIẾU TIÊU CHUẨN NĂNG LỰC', TIEUDEM)

            top_rowD = 4
            bottom_rowD = 4
            left_columnD = 0
            right_columnD = 8
            ws.write_merge(top_rowD, bottom_rowD, left_columnD, right_columnD, 'TẤT CẢ KHUNG NĂNG LỰC', TIEUDEM)


            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 1500
            ws.col(1).width = 2000
            ws.col(2).width = 10000
            ws.col(3).width = 8000
            ws.col(4).width = 1000
            ws.col(5).width = 8000
            ws.col(6).width = 2000
            ws.col(6).width = 2000
            ws.col(6).width = 2000

            columns = ['STT', 'Mã ', 'Tên năng lực', 'Chức danh', 'ID','Đơn vị', 'Mức độ quan trọng', 'Mức độ thành thạo', 'Điểm tiêu chuẩn', ]
            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = queryset.values_list('stt', 'nangluc_cv__ma_nangluc', 'nangluc_cv__name',
                                        'chucdanh_CV__Ten_Nhom_CV',
                                        'chucdanh_CV',
                                        'chucdanh_CV__don_vi__Ten_DV',
                                        'Muc_quantrong_nluc',
                                        'Muc_thanhthao_nluc', 'Diem_tieuchuan', )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)

            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )

            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'TỔNG ĐIỂM', TABLE_HEADER)
            ROW_RONG ="."

            ws.write(row_num + 1, 3, ROW_RONG, TABLE_HEADER)
            ws.write(row_num + 1, 4, ROW_RONG, TABLE_HEADER)
            ws.write(row_num + 1, 5, total_muc_qt, TABLE_HEADER)
            ws.write(row_num + 1, 6, total_muc_tt, TABLE_HEADER)
            ws.write(row_num + 1, 7, Diem_tieuchuan, TABLE_HEADER)


            ws.write(row_num + 5, 1, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese

        total_queryset = queryset.count()
        context = {'queryset': queryset_f, 'form': form}

    return render(request, 'nangluc/Khung_nangluc_list_intatca.html', context)

#-------------https://www.youtube.com/watch?v=Qc5NnpxFbBo&t=4s-------------------------------------------------------------
import json

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render

# Create your views here.
from django.views.decorators.csrf import csrf_exempt

from .models import Danhgia_nluc


#-------------------Khung năng lực---------Danhgia_nangluc_quanly.html-----------------------------------------------------




#-------------------Quản lý  update năng lực---------Danhgia_nangluc_quanly.html-----------------------------------------------------
def danhgianangluc_view(request): #Danhgia_nangluc_quanly.html
    Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    form = Danhgia_nangluc_form(request.POST or None)
    danhgia_nangluc_list = Danhgia_nluc.objects.all()[1:10]
    context = {'students': danhgia_nangluc_list, 'form': form, }

    if request.method == 'POST':
        Landanhgia_nagluc = form['Landanhgia_nagluc'].value()
        Nhanvien_dg_nangluc = form['Nhanvien_dg_nangluc'].value()
        TenNangluc_congviec = form['TenNangluc_congviec'].value()

        if (Landanhgia_nagluc != ''):
            danhgia_nangluc_list = Danhgia_nluc.objects.filter(
                Landanhgia_nagluc__icontains=form['Landanhgia_nagluc'].value(),
                Nhanvien_dg_nangluc_id=Nhanvien_dg_nangluc,
            )
    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)
    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form}

    Diem_kq = Danhgia_nluc.objects.aggregate(Count('Ketqua_danhgia'))
    from django.db.models import Sum
    total_tu_dg = (danhgia_nangluc_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (danhgia_nangluc_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*10"))['total'])
    total_chung = (danhgia_nangluc_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia"))['total'])

    total_mucqt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_quantrong_nluc',
                                                  field="TenNangluc_congviec__Muc_quantrong_nluc"))['total'])
    total_tt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_thanhthao_nluc',
                                                  field="TenNangluc_congviec.Muc_thanhthao_nluc"))['total'])
    total_diemchuan = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Diem_tieuchuan',
                                                   field="TenNangluc_congviec__Diem_tieuchuan * 2"))['total'])
    total_Diem_tu_danhgia = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tu_danhgia',
                                                   field="Diem_tu_danhgia*1"))['total'])

    total_diemcdat = (danhgia_nangluc_list.aggregate(total=Sum('Diem_dat',
                                                   field="Diem_dat"))['total'])

    if total_diemcdat and total_diemchuan:
        ketqua_nangluc = round((total_diemcdat/total_diemchuan ),2)
    else:
        ketqua_nangluc =0


    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_NL_'+'.xls'
        #responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_nangluc + '_KhungNL.xls'
        #responese['Content-Disposition'] = 'attachment; filename='+ Nhanvien_dg_nangluc + '.xls'
       # responese['Content-Disposition'] = 'attachment; filename=str(Nhanvien_dg_nangluc)'+'d.xls'
#===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        #ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        ws = wb.add_sheet('ĐG')
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)
#  --------------------------------------------
        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ NĂNG LỰC NHÂN SỰ', TIEUDEM)

        rows1 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1]
        rows2 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 1800
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'Tên năng lực', 'Mức độ quan trọng', 'Mức độ Thành thạo', 'Điểm tiêu chuẩn',' Tự đánh giá',
                   'Cấp trên đánh giá', 'Kết quả thống nhất','Điểm ĐG',  'Kết quả']
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_nangluc_list.values_list('TenNangluc_congviec__stt', 'TenNangluc_congviec__nangluc_cv__ma_nangluc',
                                                'TenNangluc_congviec__nangluc_cv__name',
                                                'TenNangluc_congviec__Muc_quantrong_nluc',
                                                'TenNangluc_congviec__Muc_thanhthao_nluc',
                                                'TenNangluc_congviec__Diem_tieuchuan',
                                                'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Diem_dat', 'Ketqua_tile')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 3, total_mucqt, TABLE_HEADER)
        ws.write(row_num + 1, 4, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 5, total_diemchuan, TABLE_HEADER)

        ws.write(row_num + 1, 6, total_tu_dg, TABLE_HEADER)# Tự đánh giá
        ws.write(row_num + 1, 7, total_ql, TABLE_HEADER) #Cấp trên đánh giá
        ws.write(row_num + 1, 8, total_chung, TABLE_HEADER) #Kết quả thống nhất

        ws.write(row_num + 1, 9, total_diemcdat, TABLE_HEADER) #Điểm đánh giá
        ws.write(row_num + 1, 10, ketqua_nangluc, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ...năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ...năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese

    if form['Xuất_Word'].value() == True:

        doc = DocxTemplate("word_template/M_danhgiangluc.docx")
        queryset5 = {
                  "Ho_ten":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1] ,
                  "Chuc_danh":(danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV'))[1] ,

                  "Don_vi":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1] ,
                  'total_diemchuan': total_diemchuan,
                  'total_diemcdat': total_diemcdat,
                  'tiledat' :ketqua_nangluc,

                   "productss": danhgia_nangluc_list,

                    }
        doc.render(queryset5)
        doc.save('thu_word/NL_'+ Nhanvien_dg_nangluc + '.docx')

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': danhgia_nangluc_list,
               "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form,
               'total_mucqt': total_mucqt, 'total_tt': total_tt,
               'total_diemchuan': total_diemchuan,

               'form': form, 'total_tu_dg': total_tu_dg,
               'total_ql': total_ql,
               'total_chung':total_chung,
               'total_Diem_tu_danhgia':total_Diem_tu_danhgia,
               'total_diemcdat': total_diemcdat,

               'ketqua_nangluc': ketqua_nangluc,


               }

    return render(request,"nangluc/Danhgia_nangluc_quanly.html",context)

@csrf_exempt#Danhgia_nangluc_aab.html
def InsertStudent(request):
    TenNangluc_congviec=request.POST.get("TenNangluc_congviec")
    Quanly_danhgia=request.POST.get("Quanly_danhgia")
    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    Ketqua_danhgia=request.POST.get("Ketqua_danhgia")

    try:
        student=Danhgia_nluc(TenNangluc_congviec=TenNangluc_congviec,Quanly_danhgia=Quanly_danhgia, tu_danhgia_dapung=tu_danhgia_dapung, Ketqua_danhgia=Ketqua_danhgia,)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)


@csrf_exempt#Danhgia_nangluc_aab.html
def update_all(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Danhgia_nluc.objects.get(id=dic_single['id'])
            #student.TenNangluc_congviec=dic_single['TenNangluc_congviec']
            student.Quanly_danhgia=dic_single['Quanly_danhgia']
            student.Ketqua_danhgia=dic_single['Ketqua_danhgia']
            #student.Diem_dat= student.Ketqua_danhgia * student.TenNangluc_congviec.Muc_quantrong_nluc

            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt#Danhgia_nangluc_aab.html
def delete_data(request):
    id=request.POST.get("id")
    try:
        student=Danhgia_nluc.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xóa không được"}
        return JsonResponse(stuent_data,safe=False)

#-------------------Kếtthúc Quản lý Đanh giá Năng lực----------------------------------------------------------

#-------------------Nhân viên update năng lực  ---------------------------------------------------------------
def danhgianangluc_view_nhanvien(request):
    Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    form = Danhgia_nangluc_form(request.POST or None)
    danhgia_nangluc_list = Danhgia_nluc.objects.all()[1:10]
    context = {'students': danhgia_nangluc_list, 'form': form, }
    if request.method == 'POST':
        Landanhgia_nagluc = form['Landanhgia_nagluc'].value()
        Nhanvien_dg_nangluc = form['Nhanvien_dg_nangluc'].value()
        TenNangluc_congviec = form['TenNangluc_congviec'].value()
        if (Landanhgia_nagluc != ''):
            danhgia_nangluc_list = Danhgia_nluc.objects.filter(
                Landanhgia_nagluc__icontains=form['Landanhgia_nagluc'].value(),
                Nhanvien_dg_nangluc_id=Nhanvien_dg_nangluc,
            )

    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)
    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form}

    Diem_kq = Danhgia_nluc.objects.aggregate(Count('Ketqua_danhgia'))
    from django.db.models import Sum
    total_tu_dg = (danhgia_nangluc_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (danhgia_nangluc_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*10"))['total'])
    total_chung = (danhgia_nangluc_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia"))['total'])

    total_mucqt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_quantrong_nluc',
                                                  field="TenNangluc_congviec__Muc_quantrong_nluc"))['total'])
    total_tt = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Muc_thanhthao_nluc',
                                                  field="TenNangluc_congviec.Muc_thanhthao_nluc"))['total'])
    total_diemchuan = (danhgia_nangluc_list.aggregate(total=Sum('TenNangluc_congviec__Diem_tieuchuan',
                                                   field="TenNangluc_congviec__Diem_tieuchuan * 2"))['total'])

    total_Diem_tu_danhgia = (danhgia_nangluc_list.aggregate(total=Sum('Diem_tu_danhgia',
                                                   field="Diem_tu_danhgia*1"))['total'])

    total_diemcdat = (danhgia_nangluc_list.aggregate(total=Sum('Diem_dat',
                                                   field="Diem_dat"))['total'])

    if  total_diemchuan  and total_Diem_tu_danhgia :
        ketqua_tu_dg_nangluc = round((total_Diem_tu_danhgia/total_diemchuan ),2)
    else:
        ketqua_tu_dg_nangluc =0



    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Tu_D_gia_NL_'+'.xls'
        #responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_nangluc + '_KhungNL.xls'
        #responese['Content-Disposition'] = 'attachment; filename='+ Nhanvien_dg_nangluc + '.xls'
       # responese['Content-Disposition'] = 'attachment; filename=str(Nhanvien_dg_nangluc)'+'d.xls'
#===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        #ws = wb.add_sheet(str(Nhanvien_dg_nangluc))
        ws = wb.add_sheet('ĐG')
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)
#  --------------------------------------------
        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU TỰ ĐÁNH GIÁ NĂNG LỰC ', TIEUDEM)

        rows1 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1]
        rows2 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1200
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 1800
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'Tên năng lực', 'Mức độ quan trọng', 'Mức độ Thành thạo', 'Điểm tiêu chuẩn',' Tự đánh giá',
                   'Điểm TỰ  ĐG',  ]
        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_nangluc_list.values_list('TenNangluc_congviec__stt', 'TenNangluc_congviec__nangluc_cv__ma_nangluc',
                                                'TenNangluc_congviec__nangluc_cv__name',
                                                'TenNangluc_congviec__Muc_quantrong_nluc',
                                                'TenNangluc_congviec__Muc_thanhthao_nluc',
                                                'TenNangluc_congviec__Diem_tieuchuan',

                                                'tu_danhgia_dapung',
                                                'Diem_tu_danhgia', )


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 160;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 2
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM:', TABLE_HEADER)

        ws.write(row_num + 1, 3, total_mucqt, TABLE_HEADER)
        ws.write(row_num + 1, 4, total_tt, TABLE_HEADER)
        ws.write(row_num + 1, 5, total_diemchuan, TABLE_HEADER)

        ws.write(row_num + 1, 6, total_tu_dg, TABLE_HEADER)
        ws.write(row_num + 1, 7, total_Diem_tu_danhgia, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ... năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese

    if form['Xuất_Word'].value() == True:

        doc = DocxTemplate("word_template/M_danhgiangluc.docx")
        queryset5 = {
                  "Ho_ten":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__ho_lot_thuong_dung')[1] ,
                  "Chuc_danh":(danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__vitri_CV__Ten_Nhom_CV'))[1] ,

                  "Don_vi":danhgia_nangluc_list.values_list('Nhanvien_dg_nangluc__bo_phan__ten_bp')[1] ,
                  'total_diemchuan': total_diemchuan,
                  'total_diemcdat': total_Diem_tu_danhgia,
                  'tiledat' :ketqua_tu_dg_nangluc,
                  'ketqua_nangluc':ketqua_tu_dg_nangluc,
                  "productss": danhgia_nangluc_list,

                    }
        doc.render(queryset5)
        doc.save('thu_word/NL_'+ Nhanvien_dg_nangluc + '.docx')

    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': danhgia_nangluc_list,
               "productss": danhgia_nangluc_list, 'students': danhgia_nangluc_list, 'form': form,
               'total_mucqt': total_mucqt, 'total_tt': total_tt,
               'total_diemchuan': total_diemchuan,
     #          'total_diemcdat': total_diemcdat,
               'form': form,
               'total_tu_dg': total_tu_dg,
               'total_ql': total_ql,
               'total_chung':total_chung,
               'total_diem_tudanhgia':total_Diem_tu_danhgia,

               'ketqua_nangluc': ketqua_tu_dg_nangluc,}
    return render(request,"nangluc/Danhgia_nangluc_nv.html",context)


@csrf_exempt
def Insert_nangluc_nv(request):
    TenNangluc_congviec=request.POST.get("TenNangluc_congviec")
    Quanly_danhgia=request.POST.get("Quanly_danhgia")
    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    try:
        student=Danhgia_nluc(TenNangluc_congviec=TenNangluc_congviec,Quanly_danhgia=Quanly_danhgia,tu_danhgia_dapung=tu_danhgia_dapung)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt
def update_all_nhanvien(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Danhgia_nluc.objects.get(id=dic_single['id'])
            #student.TenNangluc_congviec=dic_single['TenNangluc_congviec']
            student.Quanly_danhgia=dic_single['Quanly_danhgia']
            student.tu_danhgia_dapung=dic_single['tu_danhgia_dapung']
            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)
@csrf_exempt

def delete_nv_nhanvien(request):
    id=request.POST.get("id")
    try:
        student=Danhgia_nluc.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xỏa không được"}
        return JsonResponse(stuent_data,safe=False)
#-------------------Kết thúc Nhân viên update năng lực---------------------------------------------------------












