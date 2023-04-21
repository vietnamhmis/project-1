from datetime import date

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse, request
from django.db.models import Q
from django.db.models import Sum, Count
import xlwt
from django.views.generic import ListView, DetailView, View
from docxtpl import DocxTemplate, InlineImage #r

# Create your views here.
from openpyxl.styles.builtins import output
import json

from .models import  Khung_kpi, Danhgia_KPI
from mota_cv.models import Mota_Cv7
from enroll.models import Nhan_vien
from nhansu.models import KPI_list
from .forms import Mota_cv_form, Khung_KPI_form, Danhgia_KPI_form, Khung_KPI_form_update
#----------------KPI Ajax:

from django.views.generic import ListView
from .models import Khung_kpi

from django.views.generic import View
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Sum


#-------------------
def danhgiakpi_vnv(request):
    Nhanvien_dg_KPI = request.GET.get('Nhanvien_dg_KPI')
    form = Danhgia_KPI_form(request.POST or None)
    danhgia_KPI_list = Danhgia_KPI.objects.all()[1:7]
    context = {'danhgia_KPI_list1': danhgia_KPI_list, 'form': form, }

    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'students': danhgia_KPI_list, 'form': form}
    Diem_kq = Danhgia_KPI.objects.aggregate(Count('Ketqua_danhgia'))

    if request.method == 'POST':
        Landanhgia_KPI = form['Landanhgia_KPI'].value()
        Nhanvien_dg_KPI = form['Nhanvien_dg_KPI'].value()
        if (Landanhgia_KPI != ' '):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if (Nhanvien_dg_KPI != ''):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if form['Tất_cả'].value() == True:
            danhgia_KPI_list = Danhgia_KPI.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    from django.db.models import Sum
    total_tu_dg = (danhgia_KPI_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*1"))['total'])
    total_ql = (danhgia_KPI_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])

    Tile_hoanthanh = (danhgia_KPI_list.aggregate(total=Sum('Tile_hoanthanh', field="Tile_hoanthanh"))['total'])

    total_Diem_congviec = (danhgia_KPI_list.aggregate(total=Sum('Diem_congviec', field="Diem_congviec"))['total'])

    tall_Diem_trongso = (danhgia_KPI_list.aggregate(total=Sum('Diem_trongso', field="Diem_trongso"))['total'])

    total_Ketqua_tile = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_tile',field="Ketqua_tile*1"))['total'])


    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_KPI_'+'.xls'
    #===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_KPI))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ KPI', TIEUDEM)

        rows1 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__ho_lot_thuong_dung')[1]
        rows2 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 200;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1800
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 9000
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'MỤC TIÊU CHIẾN LƯỢC KPO', 'KPI', 'CHỈ TIÊU', 'ĐV tính', 'TỶ TRỌNG','TẦN SUẤT ĐO',
                   'Tự ĐG', 'Quản lý ĐG', 'KQ CHUNG','K.QUẢ THỰC HIỆN',' TỈ LỆ HOÀN THÀNH ', 'ĐIỂM CÔNG VIỆC', 'ĐIỂM TRỌNG SỐ']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_KPI_list.values_list( 'Ten_kpi__id', 'Ten_kpi__kpi_cv__ma_kpi',
                                                'Ten_kpi__kpi_cv__kpo',
                                                'Ten_kpi__kpi_cv__name',
                                                'Ten_kpi__kpi_cv__chi_tieu',
                                                'Ten_kpi__kpi_cv__donvi_tinh__name', 'Ten_kpi__ti_trong',
                                                'Ten_kpi__kpi_cv__tan_xuat_d_gia', 'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Tile_hoanthanh', 'Diem_congviec', 'Diem_trongso')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 7
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 8, total_tu_dg, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_ql, TABLE_HEADER)
        ws.write(row_num + 1, 10, total_chung, TABLE_HEADER)



        ws.write(row_num + 1, 11, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 12, Tile_hoanthanh, TABLE_HEADER)
        ws.write(row_num + 1, 13, total_Diem_congviec, TABLE_HEADER)
        ws.write(row_num + 1, 14, tall_Diem_trongso, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    chucdanh_CV2 = Nhan_vien.objects.filter()
    #context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'students': danhgia_KPI_list,'form': form,
               'Tile_hoanthanh': Tile_hoanthanh,
               'total_Diem_congviec': total_Diem_congviec,
               'tall_Diem_trongso': tall_Diem_trongso,
               'total_Ketqua_tile': total_Ketqua_tile,
               'total_tu_dg': total_tu_dg, 'total_ql': total_ql, 'total_chung':total_chung}

    return render(request, 'kpi_bsc/Danhgia_kpi_nv.html', context)
#-------------------Nhân viên update năng lực  ---------------------------------------------------------------
def danhgiakpi_vynv(request):
    Nhanvien_dg_KPI = request.GET.get('Nhanvien_dg_KPI')
    form = Danhgia_KPI_form(request.POST or None)
    queryset = Danhgia_KPI.objects.all()[1:10]
    context = {'students': queryset, 'form': form, }
    if request.method == 'POST':
        Landanhgia_KPI = form['Landanhgia_KPI'].value()
        Nhanvien_dg_KPI = form['Nhanvien_dg_KPI'].value()
        Ten_kpi = form['Ten_kpi'].value()
        if (Landanhgia_KPI != ''):
            queryset = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),
                Nhanvien_dg_KPI_id=Nhanvien_dg_KPI,
            )

    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)
    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, "productss": queryset, 'students': queryset, 'form': form}
    Diem_kq = Danhgia_KPI.objects.aggregate(Count('Ketqua_danhgia'))
    total_tu_dg = (queryset.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*3"))['total'])
    total_ql = (queryset.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*10"))['total'])
    total_chung = (queryset.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia"))['total'])
    total_mucqt = 1
    total_tt = 1
    total_diemchuan = 1
    total_Diem_tu_danhgia = 1
    total_diemcdat = 1

    if  total_diemchuan  and total_Diem_tu_danhgia :
        ketqua_tu_dg_nangluc = round((total_Diem_tu_danhgia/total_diemchuan ),2)
    else:
        ketqua_tu_dg_nangluc =0


    if form['Xuất_Excel'].value() == True:
            responese = HttpResponse(content_type='application/ms-excel')
            # responese['Content-Disposition'] = 'attachment; filename=Khung_Nang_luc'+'  '+'.xls'
            responese['Content-Disposition'] = 'attachment; filename=' + Nhanvien_dg_KPI + '_KhungKPI.xls'

            wb = xlwt.Workbook(encoding='utf-8')
            # wb.set_paper(9)
            rows1 = str(queryset.values_list('Nhanvien_dg_KPI')[1])
            if form['Tất_cả'].value() == True:
                ws = wb.add_sheet('KPI')
            else:
                ws = wb.add_sheet(rows1)


            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )
            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 5
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 5
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 5
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'KHUNG KPI THEO CHỨC DANH', TIEUDEM)



            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 1800
            ws.col(1).width = 1500
            ws.col(2).width = 15000
            ws.col(3).width = 10000
            ws.col(4).width = 3000
            ws.col(5).width = 2800
            ws.col(6).width = 2800
            ws.col(7).width = 4000

            columns = ['STT', 'Mã ', 'Tên KPO', ' TÊN KPi','Vị trí CV', 'id Vị trí','Đơn vị','ID ĐVị', 'Đơn vị tính',
                       'Tần xuất','Tỉ trọng','Chỉ tiêu',  'Loai KPI' ]

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = queryset.values_list('id', 'kpi_cv__ma_kpi','kpi_cv__kpo','kpi_cv__name',
                                        'chucdanh_CV__Ten_Nhom_CV','chucdanh_CV__id','chucdanh_CV__bo_phan__ten_bp','chucdanh_CV__bo_phan__id',
                                         'kpi_cv__donvi_tinh__name',   'kpi_cv__tan_xuat_d_gia',
                                        'ti_trong', 'kpi_cv__chi_tieu', 'kpi_cv__loai_kpi__name'  )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )

            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            TAM= ''
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'KẾT QUẢ', TABLE_HEADER)
            ws.write(row_num + 1, 3, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 4, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 5, TAM, TABLE_HEADER)


            ws.write(row_num + 5, 1, 'Ngày ... tháng .... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ....năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese




    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'danhgia_nangluc_list1': queryset,
               "productss": queryset, 'students': queryset, 'form': form,
               }
    return render(request,"kpi_bsc/Danhgia_kpi_nv.html",context)


@csrf_exempt
def Insert_kpi_nv(request):
    Ten_kpi=request.POST.get("Ten_kpi")
    Quanly_danhgia=request.POST.get("Quanly_danhgia")
    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    try:
        student=Danhgia_KPI(Ten_kpi=Ten_kpi,Quanly_danhgia=Quanly_danhgia,tu_danhgia_dapung=tu_danhgia_dapung)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt
def update_kpi_nhanvien(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Danhgia_KPI.objects.get(id=dic_single['id'])
            #student.Ten_kpi=dic_single['Ten_kpi']
            student.Quanly_danhgia=dic_single['Quanly_danhgia']
            student.tu_danhgia_dapung=dic_single['tu_danhgia_dapung']
            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)
@csrf_exempt

def delete_kpi_nhanvien(request):
    id=request.POST.get("id")
    try:
        student=Danhgia_KPI.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xỏa không được"}
        return JsonResponse(stuent_data,safe=False)
#-------------------Kết thúc Nhân viên update năng lực---------------------------------------------------------








#-------------------Khung KPI---------Khung_kpi_moi.html-----------------------------------------------------
def Khungkpi_View(request):
    chucdanh_CV = request.GET.get('chucdanh_CV')
    form = Khung_KPI_form(request.POST or None)
    Khung_kpi_list = Khung_kpi.objects.all()[1:10]
    context = {'students': Khung_kpi_list, 'form': form, }

    if request.method == 'POST':
        name = form['name'].value()
        chucdanh_CV = form['chucdanh_CV'].value()

        if (name != ''):
            Khung_kpi_list = Khung_kpi.objects.filter(
                name__icontains=form['name'].value(),
            )
        if (chucdanh_CV != ''):
            # queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV,kpi_cv_id = kpi_cv)
            Khung_kpi_list = Khung_kpi.objects.filter(chucdanh_CV_id=chucdanh_CV)

        if form['Tất_cả'].value() == True:
            Khung_kpi_list = Khung_kpi.objects.all()


    chucdanh_CV2 = Nhan_vien.objects.filter(don_vi_id=7)
    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2,
               "productss": Khung_kpi_list, 'students': Khung_kpi_list, 'form': form
               }

    from django.db.models import Sum
    total_TT = (Khung_kpi_list.aggregate(total=Sum('ti_trong', field="ti_trong"))['total'])
    total_chitieu = (Khung_kpi_list.aggregate(total=Sum('chi_tieu', field="chi_tieu"))['total'])


    if form['Xuất_Excel'].value() == True:
            responese = HttpResponse(content_type='application/ms-excel')
            responese['Content-Disposition'] = 'attachment; filename=' + chucdanh_CV + '_KhungKPI.xls'

            wb = xlwt.Workbook(encoding='utf-8')

           # rows1 = str(Khung_kpi_list.values_list('chucdanh_CV__Ten_Nhom_CV')[1])
            rows1="KPI"
            if form['Tất_cả'].value() == True:
                ws = wb.add_sheet('KPI')
            else:
                ws = wb.add_sheet(rows1)


            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )
            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 5
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 5
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 5
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'KHUNG KPI THEO CHỨC DANH', TIEUDEM)

            rows1 = Khung_kpi_list.values_list('chucdanh_CV__Ten_Nhom_CV')[1]
            rows3 = Khung_kpi_list.values_list('chucdanh_CV__bo_phan__ten_bp')[1]
            rows4 = Khung_kpi_list.values_list('name')[1]


            ws.write(5, 2, "Chức danh công việc:", TIEUDE2)
            top_row8 = 5
            bottom_row8 = 5
            left_column8 = 3
            right_column8 = 5
            ws.write_merge(top_row8, bottom_row8, left_column8, right_column8,   rows1, TIEUDE2) #cHỨC DANH

            ws.write(6, 2, "Đơn vị:", TIEUDE2)

            top_rowdv = 6
            bottom_rowdv = 6
            left_columndv = 3
            right_columndv = 5
            ws.write_merge(top_rowdv, bottom_rowdv, left_columndv, right_columndv,    rows3, TIEUDE2)
            ws.write(7, 2, "Đợt đanh giá KPI:", TIEUDE2)
            ws.write(7,3,  rows4, TIEUDE2)
            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 1800
            ws.col(1).width = 1500
            ws.col(2).width = 15000
            ws.col(3).width = 10000
            ws.col(4).width = 3000
            ws.col(5).width = 2800
            ws.col(6).width = 2800
            ws.col(7).width = 4000

            columns = ['STT', 'Mã ', 'Tên KPO', ' TÊN KPi','Vị trí CV', 'id Vị trí','Đơn vị','ID ĐVị', 'Đơn vị tính',
                       'Tần xuất','Tỉ trọng','Chỉ tiêu',  'Loai KPI' ]

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = Khung_kpi_list.values_list('id', 'kpi_cv__ma_kpi','kpi_cv__kpo','kpi_cv__name',
                                        'chucdanh_CV__Ten_Nhom_CV','chucdanh_CV__id','chucdanh_CV__bo_phan__ten_bp','chucdanh_CV__bo_phan__id',
                                         'kpi_cv__donvi_tinh__name',   'kpi_cv__tan_xuat_d_gia',
                                        'ti_trong', 'kpi_cv__chi_tieu', 'kpi_cv__loai_kpi__name'  )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )

            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            TAM= ''
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'KẾT QUẢ', TABLE_HEADER)
            ws.write(row_num + 1, 3, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 4, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 5, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 6, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 7, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 8, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 9, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 10, total_TT, TABLE_HEADER)
            ws.write(row_num + 1, 11, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 12, TAM, TABLE_HEADER)


            ws.write(row_num + 5, 1, 'Ngày ... tháng .... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ....năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese



    context = {'Nhanvien_duoc_dg_nangluc': chucdanh_CV2, 'Khung_kpi_list1': Khung_kpi_list,
               "productss": Khung_kpi_list, 'students': Khung_kpi_list, 'form': form, 'total_TT': total_TT,
                'total_chitieu':total_chitieu,

               }

    return render(request,"kpi_bsc/Khung_kpi_moi.html",context)

@csrf_exempt#
def Insertkhung_kpi(request):
    kpi_cv=request.POST.get("kpi_cv")
    #Quanly_danhgia=request.POST.get("Quanly_danhgia")

    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    ti_trong=request.POST.get("ti_trong")

    try:
        student=Khung_kpi(kpi_cv=kpi_cv, tu_danhgia_dapung=tu_danhgia_dapung, ti_trong=ti_trong,)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)


@csrf_exempt
def updatekhung_kpi(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Khung_kpi.objects.get(id=dic_single['id'])
           # student.kpi_cv=dic_single['kpi_cv']
            student.chi_tieu=dic_single['chi_tieu']
            student.ti_trong=dic_single['ti_trong']


            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt#
def delete_khungkpi(request):
    id=request.POST.get("id")
    try:
        student=Khung_kpi.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xóa không được"}
        return JsonResponse(stuent_data,safe=False)

#-------------------Kếtthúc Khunng  KPI----------------------------------------------------------


#-------------------Quản lý  update năng lực---------Danhgia_nangluc_quanly.html-----------------------------------------------------
def danhgia_kpi_view(request):
    Nhanvien_dg_KPI = request.GET.get('Nhanvien_dg_KPI')
    form = Danhgia_KPI_form(request.POST or None)
    danhgia_KPI_list = Danhgia_KPI.objects.all()[1:7]
    context = {'danhgia_KPI_list1': danhgia_KPI_list, 'form': form, }

    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'students': danhgia_KPI_list, 'form': form}
    Diem_kq = Danhgia_KPI.objects.aggregate(Count('Ketqua_danhgia'))

    if request.method == 'POST':
        Landanhgia_KPI = form['Landanhgia_KPI'].value()
        Nhanvien_dg_KPI = form['Nhanvien_dg_KPI'].value()
        if (Landanhgia_KPI != ' '):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if (Nhanvien_dg_KPI != ''):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if form['Tất_cả'].value() == True:
            danhgia_KPI_list = Danhgia_KPI.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    from django.db.models import Sum
    total_tu_dg = (danhgia_KPI_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*1"))['total'])
    total_ql = (danhgia_KPI_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])

    Tile_hoanthanh = (danhgia_KPI_list.aggregate(total=Sum('Tile_hoanthanh', field="Tile_hoanthanh"))['total'])

    total_Diem_congviec = (danhgia_KPI_list.aggregate(total=Sum('Diem_congviec', field="Diem_congviec"))['total'])

    tall_Diem_trongso = (danhgia_KPI_list.aggregate(total=Sum('Diem_trongso', field="Diem_trongso"))['total'])

    total_Ketqua_tile = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_tile',field="Ketqua_tile*1"))['total'])


    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_KPI_'+'.xls'
    #===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_KPI))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ KPI', TIEUDEM)

        rows1 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__ho_lot_thuong_dung')[1]
        rows2 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 200;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1800
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 9000
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'MỤC TIÊU CHIẾN LƯỢC KPO', 'KPI', 'CHỈ TIÊU', 'ĐV tính', 'TỶ TRỌNG','TẦN SUẤT ĐO',
                   'Tự ĐG', 'Quản lý ĐG', 'KQ CHUNG','K.QUẢ THỰC HIỆN',' TỈ LỆ HOÀN THÀNH ', 'ĐIỂM CÔNG VIỆC', 'ĐIỂM TRỌNG SỐ']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_KPI_list.values_list( 'Ten_kpi__id', 'Ten_kpi__kpi_cv__ma_kpi',
                                                'Ten_kpi__kpi_cv__kpo',
                                                'Ten_kpi__kpi_cv__name',
                                                'Ten_kpi__kpi_cv__chi_tieu',
                                                'Ten_kpi__kpi_cv__donvi_tinh__name', 'Ten_kpi__ti_trong',
                                                'Ten_kpi__kpi_cv__tan_xuat_d_gia', 'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Tile_hoanthanh', 'Diem_congviec', 'Diem_trongso')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 7
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 8, total_tu_dg, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_ql, TABLE_HEADER)
        ws.write(row_num + 1, 10, total_chung, TABLE_HEADER)



        ws.write(row_num + 1, 11, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 12, Tile_hoanthanh, TABLE_HEADER)
        ws.write(row_num + 1, 13, total_Diem_congviec, TABLE_HEADER)
        ws.write(row_num + 1, 14, tall_Diem_trongso, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    chucdanh_CV2 = Nhan_vien.objects.filter()
    #context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'students': danhgia_KPI_list,'form': form,
               'Tile_hoanthanh': Tile_hoanthanh,
               'total_Diem_congviec': total_Diem_congviec,
               'tall_Diem_trongso': tall_Diem_trongso,
               'total_Ketqua_tile': total_Ketqua_tile,
               'total_tu_dg': total_tu_dg, 'total_ql': total_ql, 'total_chung':total_chung}

    return render(request, 'kpi_bsc/Danhgia_kpi_quanly.html', context)

@csrf_exempt#Danhgia_nangluc_aab.html
def Insert_kpi(request):
    Ten_kpi=request.POST.get("Ten_kpi")
    Quanly_danhgia=request.POST.get("Quanly_danhgia")
    tu_danhgia_dapung=request.POST.get("tu_danhgia_dapung")
    Ketqua_danhgia=request.POST.get("Ketqua_danhgia")

    try:
        student=Danhgia_KPI(Ten_kpi=Ten_kpi,Quanly_danhgia=Quanly_danhgia, tu_danhgia_dapung=tu_danhgia_dapung, Ketqua_danhgia=Ketqua_danhgia,)
        student.save()
        stuent_data={"id":student.id,"start_date":student.start_date,"error":False,"errorMessage":"Thêm dữ liệu thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không thêm được"}
        return JsonResponse(stuent_data,safe=False)


@csrf_exempt#Danhgia_nangluc_aab.html
def update_kpi(request):
    data=request.POST.get("data")
    dict_data=json.loads(data)
    try:
        for dic_single in dict_data:
            student=Danhgia_KPI.objects.get(id=dic_single['id'])
            #student.Ten_kpi=dic_single['Ten_kpi']
            student.Quanly_danhgia=dic_single['Quanly_danhgia']
            student.Ketqua_danhgia=dic_single['Ketqua_danhgia']
            #student.Diem_dat= student.Ketqua_danhgia * student.Ten_kpi.Muc_quantrong_nluc

            student.save()
        stuent_data={"error":False,"errorMessage":"Cập nhật Thành công"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Không cập nhật được"}
        return JsonResponse(stuent_data,safe=False)

@csrf_exempt#Danhgia_nangluc_aab.html
def delete_kpi(request):
    id=request.POST.get("id")
    try:
        student=Danhgia_KPI.objects.get(id=id)
        student.delete()
        stuent_data={"error":False,"errorMessage":"Xóa hoàn thành"}
        return JsonResponse(stuent_data,safe=False)
    except:
        stuent_data={"error":True,"errorMessage":"Xóa không được"}
        return JsonResponse(stuent_data,safe=False)

#-------------------Kếtthúc Quản lý Đanh giá Năng lực----------------------------------------------------------
































































































#class Khungkpi_View(ListView):
   # model = Khung_kpi
  #  template_name = 'kpi_bsc/crud2.html'
   # context_object_name = 'users'
   # paginate_by = 5

class Create_khungKPI(View):

    def  get(self, request):
        id1 = request.GET.get('id', None)
        name1 = request.GET.get('name', None)
        user1 = request.GET.get('user', None)
        chucdanh_CV1 = request.GET.get('chucdanh_CV', None)
        kpi_cv1 = request.GET.get('kpi_cv', None)
        ti_trong1 = request.GET.get('ti_trong', None)

        obj = Khung_kpi.objects.create(
            name = name1,
            address = user1,
            chucdanh_CV = chucdanh_CV1,
            kpi_cv = kpi_cv1,
            ti_trong = ti_trong1
        )
        user = {'id':obj.id,'name':obj.name,'user':obj.user,'chucdanh_CV':obj.chucdanh_CV,'kpi_cv':obj.kpi_cv,'ti_trong':obj.ti_trong,}

        data = {
            'user': user,

        }
        return JsonResponse(data)

class CreateCrudUser(View):
    def  get(self, request):
        form = Khung_KPI_form_update(request.GET or None)
        obj = Khung_kpi.objects.create(
            name = form.name,
            user = form.user,
            chucdanh_CV = form.chucdanh_CV,
            kpi_cv = form.kpi_cv,
            ti_trong = form.ti_trong,
        )
        user = {'id':obj.id,'name':obj.name,'user':obj.user,'chucdanh_CV':obj.chucdanh_CV,'kpi_cv':obj.kpi_cv,'ti_trong':obj.ti_trong,}
        data = {
            'user': user,
            'form': form
        }
        return JsonResponse(data)


class Update_Khungkpi(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)

        name1 = request.GET.get('name', None)
        user1 = request.GET.get('user', None)
        age1 = request.GET.get('age', None)

        obj = Khung_kpi.objects.get(id=id1)
        obj.name = name1
        obj.user = user1
        obj.age = age1
        obj.save()
        user = {'id':obj.id,'name':obj.name,'address':obj.address,'age':obj.age}
        data = {
            'user': user
        }
        return JsonResponse(data)




class Delete_Khungkpi(View):
    def  get(self, request):
        id1 = request.GET.get('id', None)
        Khung_kpi.objects.get(id=id1).delete()
        data = {
            'deleted': True
        }
        return JsonResponse(data)






#----------------Khung KPI:::::::::::::
def Khung_kpi_list_2(request):
    form = Mota_cv_form(request.POST or None)
    chucdanh_CV = request.GET.get('chucdanh_CV')
    print("chucdanh_CV:" 'chucdanh_CV')

    if chucdanh_CV == None:
        kpi_list_2 = Khung_kpi.objects.all()[1:7]
    else:
        kpi_list_2 = Khung_kpi.objects.filter(chucdanh_CV__Ten_vitri_full=chucdanh_CV)

    chucdanh_CV7 = Mota_Cv7.objects.filter(don_vi_id=13)

    context = {'chucdanh_CV7': chucdanh_CV7, 'list_kpi': kpi_list_2, 'form': form}

    return render(request, 'kpi_bsc/Khungkpi_list_2.html', context)




@login_required
def add_khung_kpi(request):
    ten_nhom_nangluc = Mota_Cv7.objects.all()
    if request.method == 'POST':
        data = request.POST
        #kpi_cv2 = KPI_list.objects.filter(loai_kpi_id=1, loai_kpi_id=2, loai_kpi_id=3)  # Năng lực: chung=1; Toàn Công ty =2; chuyên môn =3, khác =4
        kpi_cv2 = KPI_list.objects.filter(loai_kpi_id__in=[1,2], Su_dung =True)
        Mota_Cv7s= Mota_Cv7.objects.all()
        #Mota_Cv7s = Mota_Cv7.objects.filter(Loai_laodong_id=1)
        # Mota_Cv7s= Mota_Cv7.objects.filter (Q(Loai_laodong_id=3)|Q(Loai_laodong_id=4)|Q(Loai_laodong_id=5)|Q(Loai_laodong_id=6)|Q(Loai_laodong_id=7))
        print('data:', data)
        print('Tên công việc', kpi_cv2)

        for kpi_cv in kpi_cv2:
            for chucdanh_CV in Mota_Cv7s:
                khoa_hoc = Khung_kpi.objects.create(
                    name=data['name'],
                    user=request.user,
                    kpi_cv=kpi_cv,
                    stt=data['ti_trong'],
                    chucdanh_CV=chucdanh_CV,
                )
        return redirect('Khung_kpi_list_2')
    context = {'ten_nhom_nangluc': ten_nhom_nangluc}
    return render(request, 'kpi_bsc/Khung_kpi_add.html')
#----------------
@login_required
def Khung_kpi_list(request):
    form = Khung_KPI_form(request.POST or None)
    queryset = Khung_kpi.objects.filter()[1:10]
    context = {'queryset': queryset, 'form': form, }

    if request.method == 'POST':
      #  kpi_cv = form['kpi_cv'].value()
        chucdanh_CV = form['chucdanh_CV'].value()
        name = form['name'].value()

        if (name != ''):
            queryset = Khung_kpi.objects.filter(
                name__icontains=form['name'].value(),
            )
        if (chucdanh_CV != ''):

            queryset = queryset.filter(chucdanh_CV_id=chucdanh_CV)
        if form['Tất_cả'].value() == True:
            queryset = Khung_kpi.objects.all()

        context = {'queryset': queryset, 'form': form, }


     #   total_Ti_trong = (Khung_kpi.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('ti_trong', field="ti_trong"))['total'])
      #  total_muc_tt = (Khung_kpi.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Muc_thanhthao_nluc', field="Muc_thanhthao_nluc"))['total'])
       # Diem_tieuchuan = (Khung_kpi.objects.filter(chucdanh_CV_id=chucdanh_CV).aggregate(total=Sum('Diem_tieuchuan', field="Diem_tieuchuan"))['total'])


        if form['Xuất_Excel'].value() == True:
            responese = HttpResponse(content_type='application/ms-excel')
            # responese['Content-Disposition'] = 'attachment; filename=Khung_Nang_luc'+'  '+'.xls'
            responese['Content-Disposition'] = 'attachment; filename=' + chucdanh_CV + '_KhungKPI.xls'

            wb = xlwt.Workbook(encoding='utf-8')
            # wb.set_paper(9)
            rows1 = str(queryset.values_list('chucdanh_CV__Ten_Nhom_CV')[1])
            if form['Tất_cả'].value() == True:
                ws = wb.add_sheet('KPI')
            else:
                ws = wb.add_sheet(rows1)


            font_style = xlwt.XFStyle()
            font_style.font.bold = True
            font_style.font.shadow = True

            TIEUDEM = xlwt.easyxf(
                'font: color RED, bold 1, name calibri, height 420;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;'
            )

            TIEUDE = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE1 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE2 = xlwt.easyxf(
                'font: color BLUE, bold 1, name Calibri, height 250;'
 
                'pattern:pattern_back_colour dark_red_ega;')

            TIEUDE3 = xlwt.easyxf(
                'font: color RED, bold 1, name Calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'pattern:pattern_back_colour dark_red_ega'
            )
            top_row5 = 0
            bottom_row5 = 0
            left_column5 = 0
            right_column5 = 2
            ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)

            top_row7 = 0
            bottom_row7 = 0
            left_column7 = 3
            right_column7 = 5
            ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

            top_row9 = 1
            bottom_row9 = 1
            left_column9 = 3
            right_column9 = 5
            ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

            top_row6 = 1
            bottom_row6 = 1
            left_column6 = 0
            right_column6 = 2
            ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)

            top_row = 3
            bottom_row = 3
            left_column = 0
            right_column = 5
            ws.write_merge(top_row, bottom_row, left_column, right_column, 'KHUNG KPI THEO CHỨC DANH', TIEUDEM)

            rows1 = queryset.values_list('chucdanh_CV__Ten_Nhom_CV')[1]
            rows3 = queryset.values_list('chucdanh_CV__bo_phan__ten_bp')[1]
            rows4 = queryset.values_list('name')[1]


            ws.write(5, 2, "Chức danh công việc:", TIEUDE2)
            top_row8 = 5
            bottom_row8 = 5
            left_column8 = 3
            right_column8 = 5
            ws.write_merge(top_row8, bottom_row8, left_column8, right_column8,   rows1, TIEUDE2) #cHỨC DANH

            ws.write(6, 2, "Đơn vị:", TIEUDE2)

            top_rowdv = 6
            bottom_rowdv = 6
            left_columndv = 3
            right_columndv = 5
            ws.write_merge(top_rowdv, bottom_rowdv, left_columndv, right_columndv,    rows3, TIEUDE2)
            ws.write(7, 2, "Đợt đanh giá KPI:", TIEUDE2)
            ws.write(7,3,  rows4, TIEUDE2)
            row_num = 9
            for_left = xlwt.easyxf(
                "font: bold 1, color blue; borders: top double, bottom double, left double, right double; align: horiz left")

            TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )
            TABLE_HEADER=xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 250;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: top double, bottom double, left double, right double;'
                'pattern: pattern solid, pattern_fore_colour yellow, pattern_back_colour dark_red_ega;'
            )
            # ----Định dạng chiều rộng cột
            ws.col(0).width = 1800
            ws.col(1).width = 1500
            ws.col(2).width = 15000
            ws.col(3).width = 10000
            ws.col(4).width = 3000
            ws.col(5).width = 2800
            ws.col(6).width = 2800
            ws.col(7).width = 4000

            columns = ['STT', 'Mã ', 'Tên KPO', ' TÊN KPi','Vị trí CV', 'id Vị trí','Đơn vị','ID ĐVị', 'Đơn vị tính',
                       'Tần xuất','Tỉ trọng','Chỉ tiêu',  'Loai KPI' ]

            for col_num in range(len(columns)):
                ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
            # ----Định dạng chữ
            stt = 1
            font_style = xlwt.XFStyle()
            font_style.font.italic = False
            for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

            rows = queryset.values_list('id', 'kpi_cv__ma_kpi','kpi_cv__kpo','kpi_cv__name',
                                        'chucdanh_CV__Ten_Nhom_CV','chucdanh_CV__id','chucdanh_CV__bo_phan__ten_bp','chucdanh_CV__bo_phan__id',
                                         'kpi_cv__donvi_tinh__name',   'kpi_cv__tan_xuat_d_gia',
                                        'ti_trong', 'kpi_cv__chi_tieu', 'kpi_cv__loai_kpi__name'  )
            for row in rows:
                row_num += 1
                for col_num in range(len(row)):
                    if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                    else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


            for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
            for_foot_ng = xlwt.easyxf('font: color blue, name calibri, height 220;'
                                      'align: vertical center, horizontal center,')
            TABLE_HEADER2 = xlwt.easyxf(
                'font: bold 1, color blue, name calibri, height 220;'
                'align: vertical center, horizontal center;'
            )

            top_row2 = row_num + 1
            bottom_row2 = row_num + 1
            left_column2 = 0
            right_column2 = 2
            TAM= ''
            ws.write_merge(top_row2, bottom_row2, left_column2, right_column2, 'KẾT QUẢ', TABLE_HEADER)
            ws.write(row_num + 1, 3, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 4, TAM, TABLE_HEADER)
            ws.write(row_num + 1, 5, TAM, TABLE_HEADER)


            ws.write(row_num + 5, 1, 'Ngày ... tháng .... năm 2023', for_foot_ng)
            ws.write(row_num + 6, 1, 'Người phê duyệt', TABLE_HEADER2)
            ws.write(row_num + 5, 3, 'Ngày .... tháng ....năm 2023', for_foot_ng)
            ws.write(row_num + 6, 3, 'Người xem xét', TABLE_HEADER2)
            ws.write(row_num + 5, 5, 'Ngày ... tháng ... năm 2023', for_foot_ng)
            # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
            ws.write(row_num + 6, 5, 'Người thiết lập', TABLE_HEADER2)

            wb.save(responese)
            return responese

        total_queryset = queryset.count()
        context = {'queryset': queryset, 'form': form}
    return render(request, 'kpi_bsc/Khung_kpi_list.html', context)


def update_Khung_KPI(request, id):
    if request.method == 'POST':
        dv = Khung_kpi.objects.get(pk=id)
        fmdv = Khung_KPI_form_update(request.POST, instance=dv)
        if fmdv.is_valid():
            fmdv.save()
            messages.info(request, "KPI đã lưu")
    else:
        dv = Khung_kpi.objects.get(pk=id)

        fmdv = Khung_KPI_form_update(instance=dv)
    return render(request, 'kpi_bsc/Khung_KPI_up.html', {'form': fmdv})


def del_Khung_KPI(request, id):
    dv = Khung_kpi.objects.get(pk=id)
    if request.method == 'POST':
        dv.delete()
      #  return HttpResponseRedirect('/cvnl/')
    return render(request, 'kpi_bsc/Khung_KPI_delete.html')



# Đánh giá  KPI:::::::::::::::::::::::::
def danhgia_KPI_list(request):

    Nhanvien_dg_KPI = request.GET.get('Nhanvien_dg_KPI')
    form = Danhgia_KPI_form(request.POST or None)
    danhgia_KPI_list = Danhgia_KPI.objects.all()[1:7]
    context = {'danhgia_KPI_list1': danhgia_KPI_list, 'form': form, }

    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_kpi_list2': danhgia_KPI_list, 'form': form}
    Diem_kq = Danhgia_KPI.objects.aggregate(Count('Ketqua_danhgia'))

    if request.method == 'POST':
        Landanhgia_KPI = form['Landanhgia_KPI'].value()
        Nhanvien_dg_KPI = form['Nhanvien_dg_KPI'].value()
        if (Landanhgia_KPI != ' '):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if (Nhanvien_dg_KPI != ''):
            danhgia_KPI_list = Danhgia_KPI.objects.filter(
                Landanhgia_KPI__icontains=form['Landanhgia_KPI'].value(),Nhanvien_dg_KPI_id=Nhanvien_dg_KPI
            )
        if form['Tất_cả'].value() == True:
            danhgia_KPI_list = Danhgia_KPI.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    from django.db.models import Sum
    total_tu_dg = (danhgia_KPI_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*1"))['total'])
    total_ql = (danhgia_KPI_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])

    Tile_hoanthanh = (danhgia_KPI_list.aggregate(total=Sum('Tile_hoanthanh', field="Tile_hoanthanh"))['total'])

    total_Diem_congviec = (danhgia_KPI_list.aggregate(total=Sum('Diem_congviec', field="Diem_congviec"))['total'])

    tall_Diem_trongso = (danhgia_KPI_list.aggregate(total=Sum('Diem_trongso', field="Diem_trongso"))['total'])

    total_Ketqua_tile = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_tile',field="Ketqua_tile*1"))['total'])


    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_KPI_'+'.xls'
    #===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_KPI))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ KPI', TIEUDEM)

        rows1 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__ho_lot_thuong_dung')[1]
        rows2 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 200;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1800
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 9000
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'MỤC TIÊU CHIẾN LƯỢC KPO', 'KPI', 'CHỈ TIÊU', 'ĐV tính', 'TỶ TRỌNG','TẦN SUẤT ĐO',
                   'Tự ĐG', 'Quản lý ĐG', 'KQ CHUNG','K.QUẢ THỰC HIỆN',' TỈ LỆ HOÀN THÀNH ', 'ĐIỂM CÔNG VIỆC', 'ĐIỂM TRỌNG SỐ']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_KPI_list.values_list( 'Ten_kpi__id', 'Ten_kpi__kpi_cv__ma_kpi',
                                                'Ten_kpi__kpi_cv__kpo',
                                                'Ten_kpi__kpi_cv__name',
                                                'Ten_kpi__kpi_cv__chi_tieu',
                                                'Ten_kpi__kpi_cv__donvi_tinh__name', 'Ten_kpi__ti_trong',
                                                'Ten_kpi__kpi_cv__tan_xuat_d_gia',
                                                'Ten_kpi__kpi_cv__chi_tieu', 'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Tile_hoanthanh', 'Diem_congviec', 'Diem_trongso')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 7
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 8, total_tu_dg, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_ql, TABLE_HEADER)
        ws.write(row_num + 1, 10, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 11, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 12, Tile_hoanthanh, TABLE_HEADER)
        ws.write(row_num + 1, 13, total_Diem_congviec, TABLE_HEADER)
        ws.write(row_num + 1, 14, tall_Diem_trongso, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    chucdanh_CV2 = Nhan_vien.objects.filter()
    #context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list, 'form': form,}

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_KPI_list1': danhgia_KPI_list,'form': form,
               'Tile_hoanthanh': Tile_hoanthanh,
               'total_Diem_congviec': total_Diem_congviec,
               'tall_Diem_trongso': tall_Diem_trongso,
               'total_Ketqua_tile': total_Ketqua_tile,
               'total_tu_dg': total_tu_dg, 'total_ql': total_ql, 'total_chung':total_chung}

    return render(request, 'kpi_bsc/Danhgia_kpi_chitiet1.html', context)


@login_required
def add_danhgia_kpi(request):
    # Nhanvien_dg_nangluc = request.GET.get('Nhanvien_dg_nangluc')
    Nhanvien_danhgia = Nhan_vien.objects.all()
    if request.method == 'POST':
        data = request.POST
        # Nhan_viens = Nhan_vien.objects.filter(Q(vitri_CV_id=1)|Q(vitri_CV_id=2))
        Nhan_viens = Nhan_vien.objects.filter()

        Congviec_nanglucs = Khung_kpi.objects.filter()
        # Congviec_nanglucs = Congviec_nangluc.objects.all()

     #   print('Nhập dự liệu ở form:', data)
    #    print('Tên Nhân viên', Nhan_viens)
        for Nhanvien_dg_nangluc in Nhan_viens:
            for Ten_congviec in Congviec_nanglucs:
                if Nhanvien_dg_nangluc.vitri_CV_id == Ten_congviec.chucdanh_CV_id:
                    khoa_hoc = Danhgia_KPI.objects.create(
                        Landanhgia_KPI=data['Landanhgia_KPI'],
                        Nhanvien_dg_KPI=Nhanvien_dg_nangluc,
                        Ten_kpi=Ten_congviec,

                        )
                else:
                    continue
        return redirect('/')
    context = {'Nhanvien_danhgias': Nhanvien_danhgia}

    return render(request, 'kpi_bsc/Danhgia_kpi_add.html')


@login_required
#Danhgia_nangluc2.htm
def Danhgia_kpi_list2(request):
    Nhanvien_dg_KPI = request.GET.get('Nhanvien_dg_kpi')
    form = Danhgia_KPI_form(request.POST or None)

    if Nhanvien_dg_KPI == None:
        danhgia_KPI_list = Danhgia_KPI.objects.all()[1:50]
    else:

        danhgia_KPI_list = Danhgia_KPI.objects.filter(Nhanvien_dg_KPI__ho_lot_thuong_dung=Nhanvien_dg_KPI)

    chucdanh_CV2 = Nhan_vien.objects.all()

    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_kpi_list2': danhgia_KPI_list, 'form': form}
    Diem_kq = Danhgia_KPI.objects.aggregate(Count('Ketqua_danhgia'))

    from django.db.models import Sum
    total_tu_dg = (danhgia_KPI_list.aggregate(total=Sum('tu_danhgia_dapung', field="tu_danhgia_dapung*1"))['total'])
    total_ql = (danhgia_KPI_list.aggregate(total=Sum('Quanly_danhgia', field="Quanly_danhgia*1"))['total'])
    total_chung = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_danhgia', field="Ketqua_danhgia*1"))['total'])

    Tile_hoanthanh = (danhgia_KPI_list.aggregate(total=Sum('Tile_hoanthanh', field="Tile_hoanthanh"))['total'])

    total_Diem_congviec = (danhgia_KPI_list.aggregate(total=Sum('Diem_congviec', field="Diem_congviec"))['total'])

    tall_Diem_trongso = (danhgia_KPI_list.aggregate(total=Sum('Diem_trongso', field="Diem_trongso"))['total'])

    total_Ketqua_tile = (danhgia_KPI_list.aggregate(total=Sum('Ketqua_tile',field="Ketqua_tile*1"))['total'])

    if form['Xuất_Excel'].value() == True:
        responese = HttpResponse(content_type='application/ms-excel')
        responese['Content-Disposition'] = 'attachment; filename=Danh_gia_KPI_'+'.xls'
    #===============================================================================================================
        wb = xlwt.Workbook(encoding='utf-8')
        ws = wb.add_sheet(str(Nhanvien_dg_KPI))
        font_style = xlwt.XFStyle()
        font_style.font.bold = True
        font_style.font.shadow = True
  #----------------------------------

        TIEUDEM = xlwt.easyxf(
                    'font: color RED, bold 1, name calibri, height 420;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;'
                )

        TIEUDE = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE1 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 280;'
                    'align: horiz right, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE2 = xlwt.easyxf(
                    'font: color BLUE, bold 1, name Calibri, height 280;'
     
                    'pattern:pattern_back_colour dark_red_ega;')

        TIEUDE3 = xlwt.easyxf(
                    'font: color RED, bold 1, name Calibri, height 250;'
                    'align: vertical center, horizontal center, wrap on;'
                    'pattern:pattern_back_colour dark_red_ega'
                )


        top_row5 = 0
        bottom_row5 = 0
        left_column5 = 0
        right_column5 = 2
        ws.write_merge(top_row5, bottom_row5, left_column5, right_column5,  'UBND TP.HỒ CHÍ MINH',TIEUDE)
        top_row7 = 0
        bottom_row7 = 0
        left_column7 = 5
        right_column7 = 9
        ws.write_merge(top_row7, bottom_row7, left_column7, right_column7,  'CỘNG HÒA XÃ HỘI CHỦ NGHĨA VIỆT NAM', TIEUDE)

        top_row9 = 1
        bottom_row9 = 1
        left_column9 = 5
        right_column9 = 9
        ws.write_merge(top_row9, bottom_row9, left_column9, right_column9, 'Độc lập-Tự do-Hạnh phúc', TIEUDE3)

        top_row6 = 1
        bottom_row6 = 1
        left_column6 = 0
        right_column6 = 2
        ws.write_merge(top_row6, bottom_row6, left_column6, right_column6,  'TỔNG CÔNG TY TM SÀI GÒN-TNHH MTV', TIEUDE3)


#  --------------------------------------------

        top_row = 4
        bottom_row = 4
        left_column = 0
        right_column = 9
        ws.write_merge(top_row, bottom_row, left_column, right_column, 'BIỂU ĐÁNH GIÁ KPI', TIEUDEM)

        rows1 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__ho_lot_thuong_dung')[1]
        rows2 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__vitri_CV__Ten_Nhom_CV')[1]
        rows3 = danhgia_KPI_list.values_list('Nhanvien_dg_KPI__bo_phan__ten_bp')[1]
        row_phay = (", ",)

        ws.write(6, 2, "Họ tên:", TIEUDE1)
        ws.write(6, 3, rows1 +row_phay+ rows2, TIEUDE2)
        ws.write(7, 2, "Đơn vị:", TIEUDE1)
        ws.write(7, 3, rows3, TIEUDE2)

        row_num = 9

        TABLE_HEADER = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center, wrap on;'
            'borders: top double, bottom double, left double, right double;'
            'pattern: pattern solid, pattern_fore_colour light_yellow, pattern_back_colour dark_red_ega;'
        )
        for_left = xlwt.easyxf(
                "font: color blue, name calibri, height 250; borders: left thin, right thin, top thin, bottom thin; pattern: pattern solid, fore_color white;")

        TABLE_row=xlwt.easyxf(
                'font: bold 1, color white, name calibri, height 200;'
                'align: vertical center, horizontal center, wrap on;'
                'borders: left thin, right thin, top thin, bottom thin;'
                'pattern: pattern solid, pattern_fore_colour green;'
            )

        ws.col(0).width = 1800
        ws.col(1).width = 2000
        ws.col(2).width = 8500
        ws.col(3).width = 9000
        ws.col(4).width = 2000
        ws.col(5).width = 1900
        ws.col(6).width = 1900
        ws.col(7).width = 2000
        ws.col(8).width = 2000
        ws.col(9).width = 1800
        # ------------------
        columns = ['Số TT', 'Mã NL', 'MỤC TIÊU CHIẾN LƯỢC KPO', 'KPI', 'CHỈ TIÊU', 'ĐV tính', 'TỶ TRỌNG','TẦN SUẤT ĐO',
                   'Tự ĐG', 'Quản lý ĐG', 'KQ CHUNG','K.QUẢ THỰC HIỆN',' TỈ LỆ HOÀN THÀNH ', 'ĐIỂM CÔNG VIỆC', 'ĐIỂM TRỌNG SỐ']

        for col_num in range(len(columns)):
            ws.write(row_num, col_num, columns[col_num], TABLE_HEADER)
        # ----Định dạng chữ
        font_style = xlwt.XFStyle()
        font_style.font.italic = False


        rows = danhgia_KPI_list.values_list( 'Ten_kpi__stt', 'Ten_kpi__kpi_cv__ma_kpi',
                                                'Ten_kpi__kpi_cv__kpo',
                                                'Ten_kpi__kpi_cv__name',
                                                'Ten_kpi__kpi_cv__chi_tieu',
                                                'Ten_kpi__kpi_cv__donvi_tinh__name', 'Ten_kpi__ti_trong',
                                                'Ten_kpi__kpi_cv__tan_xuat_d_gia',
                                                'Ten_kpi__kpi_cv__chi_tieu', 'tu_danhgia_dapung', 'Quanly_danhgia', 'Ketqua_danhgia',
                                                'Tile_hoanthanh', 'Diem_congviec', 'Diem_trongso')


        for row in rows:
            row_num += 1
            for col_num in range(len(row)):
                if (row[col_num]) != 0 and (row[col_num]) != 'I.NĂNG LỰC CHUNG' and (row[col_num]) != "II.NĂNG LỰC QUẢN LÝ" and (row[col_num]) != "III.NĂNG LỰC CHUYÊN MÔN"\
                            and (row[col_num]) != 'CH00' and (row[col_num]) != "QL00" and (row[col_num]) != "CM00":
                        ws.write(row_num, col_num, (row[col_num]),for_left)
                else:
                        ws.write(row_num, col_num, (row[col_num]),TABLE_row)


        print('Số cột', col_num)
        print('Số hàng', row_num)
        for_foot = xlwt.easyxf("font: color blue;  pattern: pattern solid, fore_color white;")
        for_foot_ng = xlwt.easyxf('font: color blue, name Tahoma, height 220;'
                                  'align: vertical center, horizontal center,')
        TABLE_HEADER2 = xlwt.easyxf(
            'font: bold 1, color blue, name Tahoma, height 220;'
            'align: vertical center, horizontal center;'
        )
        tam1=""

        top_row2 = row_num + 1
        bottom_row2 = row_num + 1
        left_column2 = 0
        right_column2 = 7
        ws.write_merge(top_row2, bottom_row2, left_column2, right_column2,  'TỔNG ĐIỂM', TABLE_HEADER)

        ws.write(row_num + 1, 8, total_tu_dg, TABLE_HEADER)
        ws.write(row_num + 1, 9, total_ql, TABLE_HEADER)
        ws.write(row_num + 1, 10, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 11, total_chung, TABLE_HEADER)
        ws.write(row_num + 1, 12, Tile_hoanthanh, TABLE_HEADER)
        ws.write(row_num + 1, 13, total_Diem_congviec, TABLE_HEADER)
        ws.write(row_num + 1, 14, tall_Diem_trongso, TABLE_HEADER)

        ws.write(row_num + 3, 1, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 1, 'Người phê duyệt', TABLE_HEADER2)
        ws.write(row_num + 3, 3, 'Ngày .... tháng ... năm 2023', for_foot_ng)
        ws.write(row_num + 4, 3, 'Người đánh giá', TABLE_HEADER2)
        ws.write(row_num + 3, 7, 'Ngày ... tháng ....năm 2023', for_foot_ng)
        # ws.write_merge(row_num + 18,row_num + 19, row_num + 20,row_num + 21,  'Người được đánh giá', TABLE_HEADER2)
        ws.write(row_num + 4, 7, 'Người được đánh giá', TABLE_HEADER2)
        wb.save(responese)
        return responese


    context = {'Nhanvien_duoc_dg_kpi': chucdanh_CV2, 'danhgia_kpi_list2': danhgia_KPI_list,
              'Tile_hoanthanh': Tile_hoanthanh,
               'total_Diem_congviec': total_Diem_congviec,
              'tall_Diem_trongso': tall_Diem_trongso,
              'total_Ketqua_tile': total_Ketqua_tile,
            #   'ketqua_nangluc': ketqua_nangluc,
               'form': form, 'total_tu_dg': total_tu_dg, 'total_ql': total_ql, 'total_chung':total_chung}
    return render(request, 'kpi_bsc/Danhgia_kpi2.html', context)


def Danhgia_KPI_update(request, id):
    if request.method == 'POST':
        dv = Danhgia_KPI.objects.get(pk=id)
        fmdg = Danhgia_KPI_form(request.POST, instance=dv)
        if fmdg.is_valid():
            fmdg.save()
            messages.success(request, 'Dữ liệu cập nhật')
    else:
        dv = Danhgia_KPI.objects.get(pk=id)
        fmdg = Danhgia_KPI_form(instance=dv)
    return render(request, 'kpi_bsc/Danhgia_KPI_up.html', {'form': fmdg})

def del_Danh_gia_KPI(request, id):
    dv = Danhgia_KPI.objects.get(pk=id)
    if request.method == 'POST':
        dv.delete()
      #  return HttpResponseRedirect('/cvnl/')
    return render(request, 'kpi_bsc/Khung_KPI_delete.html')
